#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
履约义务查询接口
功能：条款自动分类、义务字段提取、标准产品义务比对、差异检测与标注
"""

import re
import yaml
import json
from typing import Dict, List, Any, Optional, Tuple
from dataclasses import dataclass, asdict
from enum import Enum
from pathlib import Path


class ObligationType(Enum):
    """义务类型枚举"""
    DELIVERY = "OBJ-01"
    ACCEPTANCE = "OBJ-02"
    PAYMENT = "OBJ-03"
    WARRANTY = "OBJ-04"
    TECH_SUPPORT = "OBJ-05"
    TRAINING = "OBJ-06"
    INTELLECTUAL_PROPERTY = "OBJ-07"
    BREACH = "OBJ-08"


class RiskLevel(Enum):
    """风险等级枚举"""
    CRITICAL = "critical"
    HIGH = "high"
    MEDIUM = "medium"
    LOW = "low"


@dataclass
class ExtractedField:
    """提取的字段"""
    name: str
    value: str
    confidence: float
    source: str


@dataclass
class Obligation:
    """履约义务对象"""
    obligation_id: str
    obligation_type: str
    category_name: str
    subcategory: str
    content_description: str
    source_text: str
    trigger_conditions: List[str]
    deadline: Optional[str] = None
    location: Optional[str] = None
    responsible_party: str = "乙方"
    beneficiary_party: str = "甲方"
    extracted_fields: Dict[str, Any] = None
    payment_related: Dict[str, Any] = None
    risk_level: str = "medium"
    risk_factors: List[str] = None
    standard_comparison: Dict[str, Any] = None
    confidence: float = 0.0
    notes: str = ""

    def __post_init__(self):
        if self.extracted_fields is None:
            self.extracted_fields = {}
        if self.risk_factors is None:
            self.risk_factors = []
        if self.trigger_conditions is None:
            self.trigger_conditions = []


class ObligationQuery:
    """履约义务查询类"""

    def __init__(self, rules_path: Optional[str] = None):
        """初始化查询器"""
        if rules_path is None:
            base_path = Path(__file__).parent.parent
            rules_path = base_path / "data" / "obligation-rules" / "obligation-split-rules.yaml"
        
        self.rules_path = Path(rules_path)
        self.rules = self._load_rules()
        self.obligation_counter = {t.value: 0 for t in ObligationType}
        
    def _load_rules(self) -> Dict:
        """加载规则库"""
        with open(self.rules_path, 'r', encoding='utf-8') as f:
            return yaml.safe_load(f)
    
    def classify_clause(self, clause_text: str) -> List[Tuple[ObligationType, float]]:
        """
        条款自动分类
        返回可能的义务类型及置信度
        """
        results = []
        clause_lower = clause_text.lower()
        
        # 分类边界规则：定义每类义务的核心特征，减少交叉误判
        boundary_rules = {
            'OBJ-03':  # 付款义务 - 必须包含金额/百分比等核心特征
                {'must_have': ['%', '元', '万', '付款', '支付', '预付款', '质保金', '进度款'],
                 'boost': ['支付', '付款', '预付款', '进度款', '质保金', '违约金'],
                 'block': ['交付', '验收', '保修']},  # 关键词冲突时优先付款
            'OBJ-02':  # 验收义务 - 必须包含验收相关核心词
                {'must_have': ['验收', '检验', '检测', '不合格', '合格', '整改'],
                 'boost': ['验收', '检验', '验收标准', '验收流程', '验收合格'],
                 'block': ['支付', '付款', '违约金', '赔偿']},  # 关键词冲突时优先验收
            'OBJ-08':  # 违约责任 - 必须包含违约核心词
                {'must_have': ['违约', '逾期', '赔偿', '违约金', '解除合同', '责任'],
                 'boost': ['违约', '违约金', '赔偿', '逾期', '解除合同']},
            'OBJ-04':  # 维保义务 - 必须包含维保核心词
                {'must_have': ['保修', '维保', '维修', '质保期', '保修期'],
                 'boost': ['保修', '维保', '质保期', '保修期']},
            'OBJ-07':  # 知识产权 - 必须包含IP核心词
                {'must_have': ['知识产权', '专利', '著作权', '版权', '归属', '所有'],
                 'boost': ['知识产权', '归属', '所有', '专利']},
        }
        
        # 互斥规则：同一条款同时命中多个分类时的优先级
        priority_order = {'OBJ-08': 5, 'OBJ-03': 4, 'OBJ-02': 3, 'OBJ-04': 2, 'default': 1}
        
        # 违约责任互斥：包含违约关键词时，其他分类降权
        has_violation = any(v in clause_text for v in ['违约', '违约金', '赔偿', '逾期'])
        has_payment = any(p in clause_text for p in ['支付', '付款', '%', '元', '万'])
        has_acceptance = any(a in clause_text for a in ['验收', '检验'])
        
        for obl_type in ObligationType:
            type_code = obl_type.value
            rules = self.rules['recognition_rules'].get(type_code, {})
            keywords = rules.get('keywords', {})
            
            # 分类边界检查：有must_have规则的必须命中至少1个核心词
            if type_code in boundary_rules:
                must_haves = boundary_rules[type_code]['must_have']
                has_core = any(mh in clause_text for mh in must_haves)
                if not has_core:
                    continue  # 未命中核心特征，跳过该分类
            
            score = 0.0
            matched_high = 0
            matched_medium = 0
            
            # 高权重关键词
            for kw in keywords.get('high_weight', []):
                if kw in clause_text:
                    score += 5
                    matched_high += 1
            
            # 中权重关键词
            for kw in keywords.get('medium_weight', []):
                if kw in clause_text:
                    score += 2
                    matched_medium += 1
            
            # 特征句式匹配
            patterns = rules.get('feature_patterns', [])
            matched_patterns = 0
            for pattern in patterns:
                if self._pattern_match(pattern, clause_text):
                    score += 10
                    matched_patterns += 1
            
            # 分类边界加成：命中该分类核心特征的大幅加分
            if type_code in boundary_rules:
                boosts = boundary_rules[type_code]['boost']
                for boost_kw in boosts:
                    if boost_kw in clause_text:
                        score += 15  # 核心特征大幅加分，减少交叉误判
                
                # 互斥排除：命中block关键词的减少得分（避免交叉误判）
                blocks = boundary_rules[type_code].get('block', [])
                for block_kw in blocks:
                    if block_kw in clause_text:
                        score -= 10  # 命中排除词，降权
            
            # 优先级加成：同时命中多分类时，高优先级分类得分加成
            score += priority_order.get(type_code, 1) * 3
            
            # 计算置信度
            if score > 0:
                max_high = len(keywords.get('high_weight', []))
                max_medium = len(keywords.get('medium_weight', []))
                max_possible = 5 * max_high + 2 * max_medium + 10 * len(patterns) + 60  # 加上边界加分可能
                confidence = min(score / max(1, max_possible) * 100, 100)
                
                # 关键词匹配数量加成（合理范围）
                if matched_high >= 1:
                    confidence = min(confidence + 15, 100)
                if matched_high >= 2:
                    confidence = min(confidence + 10, 100)
                if matched_medium >= 3:
                    confidence = min(confidence + 10, 100)
                if matched_patterns > 0:
                    confidence = min(confidence + 20, 100)
                
                results.append((obl_type, confidence))
        
        # 按置信度排序
        results.sort(key=lambda x: x[1], reverse=True)
        return results
    
    def _pattern_match(self, pattern: str, text: str) -> bool:
        """简单的模式匹配"""
        # 将中文标点转换为统一格式
        text = text.replace('，', ',').replace('。', '.').replace('；', ';')
        
        # 简单的包含匹配（未来可扩展为正则）
        pattern_parts = pattern.split('.*')
        pos = 0
        for part in pattern_parts:
            if not part:
                continue
            idx = text.find(part, pos)
            if idx == -1:
                return False
            pos = idx + len(part)
        return True
    
    def extract_fields(self, clause_text: str, obl_type: ObligationType) -> Dict[str, ExtractedField]:
        """
        提取义务字段
        """
        fields = {}
        
        # 通用字段提取
        # 1. 日期时间提取
        datetime_rules = self.rules['field_extraction_regex']['datetime_extraction']
        for rule in datetime_rules:
            matches = re.findall(rule['pattern'], clause_text)
            if matches:
                field_name = rule['name']
                value = matches[0] if isinstance(matches[0], str) else ''.join(matches[0])
                fields[field_name] = ExtractedField(
                    name=field_name,
                    value=value,
                    confidence=90.0,
                    source="regex"
                )
        
        # 2. 金额/比例提取
        amount_rules = self.rules['field_extraction_regex']['amount_ratio_extraction']
        for rule in amount_rules:
            matches = re.findall(rule['pattern'], clause_text)
            if matches:
                field_name = rule['name']
                match = matches[0]
                if isinstance(match, tuple):
                    value = ''.join(m for m in match if m)
                else:
                    value = match
                fields[field_name] = ExtractedField(
                    name=field_name,
                    value=value,
                    confidence=85.0,
                    source="regex"
                )
        
        # 3. 期限提取
        period_rules = self.rules['field_extraction_regex']['period_extraction']
        for rule in period_rules:
            matches = re.findall(rule['pattern'], clause_text)
            if matches:
                field_name = rule['name']
                value = matches[0] if isinstance(matches[0], str) else ''.join(matches[0])
                fields[field_name] = ExtractedField(
                    name=field_name,
                    value=value,
                    confidence=80.0,
                    source="regex"
                )
        
        # 4. 地点提取
        location_rules = self.rules['field_extraction_regex']['location_extraction']
        for rule in location_rules:
            matches = re.findall(rule['pattern'], clause_text)
            if matches:
                field_name = rule['name']
                value = matches[0] if isinstance(matches[0], str) else ''.join(matches[0])
                fields[field_name] = ExtractedField(
                    name=field_name,
                    value=value.strip(),
                    confidence=75.0,
                    source="regex"
                )
        
        # 5. 责任方提取
        party_rules = self.rules['field_extraction_regex']['party_extraction']
        for rule in party_rules:
            if re.search(rule['pattern'], clause_text):
                party_name = rule['name']
                # 确保责任方准确：甲方/乙方分别识别
                if '甲方应' in clause_text or '甲方须' in clause_text or '甲方负责' in clause_text:
                    party_name = '甲方'
                elif '乙方应' in clause_text or '乙方须' in clause_text or '乙方负责' in clause_text:
                    party_name = '乙方'
                elif '双方应' in clause_text or '双方共同' in clause_text:
                    party_name = '双方'
                fields[party_name] = ExtractedField(
                    name=party_name,
                    value=party_name,
                    confidence=95.0,
                    source="keyword"
                )
        
        return fields
    
    def extract_responsbile_party(self, clause_text: str, obl_type: str = None) -> str:
        """提取责任方 - 准确识别甲乙双方"""
        # 违约责任特殊处理
        if obl_type == 'OBJ-08':
            if '乙方逾期' in clause_text or '乙方违约' in clause_text or '乙方未' in clause_text:
                return "乙方"
            if '甲方逾期' in clause_text or '甲方违约' in clause_text:
                return "甲方"
        
        # 优先识别甲方（付款、验收等通常是甲方义务）
        if '甲方应' in clause_text or '甲方须' in clause_text or '甲方负责' in clause_text:
            return "甲方"
        # 然后识别乙方
        elif '乙方应' in clause_text or '乙方须' in clause_text or '乙方负责' in clause_text:
            return "乙方"
        # 双方共同义务
        elif '双方应' in clause_text or '双方共同' in clause_text:
            return "双方"
        # 付款义务默认甲方
        elif obl_type == 'OBJ-03' or '支付' in clause_text or '付款' in clause_text or '预付款' in clause_text:
            return "甲方"
        # 违约默认乙方
        elif obl_type == 'OBJ-08':
            return "乙方"
        # 默认乙方
        else:
            return "乙方"
    
    def extract_deadline(self, clause_text: str) -> Optional[str]:
        """提取履行时限"""
        # 绝对日期
        date_patterns = [
            r'(\d{4}-\d{2}-\d{2})',
            r'(\d{4}年\d{1,2}月\d{1,2}日)',
            r'在(\d{4})年(\d{1,2})月(\d{1,2})日前',
        ]
        
        for pattern in date_patterns:
            match = re.search(pattern, clause_text)
            if match:
                return match.group(0)
        
        # 相对日期
        relative_patterns = [
            r'(\d+)个?工作日',
            r'(\d+)个?自然?日',
            r'(\d+)个?月',
            r'(\d+)天',
        ]
        
        for pattern in relative_patterns:
            match = re.search(pattern, clause_text)
            if match:
                return f"{match.group(0)}内"
        
        return None
    
    def determine_risk_level(self, obligation: Obligation, fields: Dict[str, ExtractedField]) -> str:
        """
        确定风险等级
        """
        risk_score = 0
        risk_factors = []
        
        # 根据义务类型基础风险
        type_code = obligation.obligation_type
        category = self.rules['obligation_categories'].get(type_code, {})
        base_risk = category.get('risk_level', 'medium')
        
        risk_map = {'low': 1, 'medium': 2, 'high': 3, 'critical': 4}
        risk_score = risk_map.get(base_risk, 2)
        
        # 检查是否有明确的履行时限
        if not obligation.deadline or obligation.deadline == "未提取":
            risk_score += 1
            risk_factors.append("无明确履行时限")
        
        # 付款义务额外风险检查
        if type_code == 'OBJ-03':
            # 预付款过高
            if '百分比' in fields:
                try:
                    ratio = float(fields['百分比'].value.replace('%', ''))
                    if ratio > 50:
                        risk_score += 2
                    elif ratio > 30:
                        risk_score += 1
                except:
                    pass
            
            # 无质保金
            if '质保金' not in obligation.content_description:
                risk_score += 1
        
        # 验收义务风险
        elif type_code == 'OBJ-02':
            if '验收标准' not in obligation.content_description:
                risk_score += 1
        
        # 知识产权风险
        elif type_code == 'OBJ-07':
            if '归属' not in obligation.content_description:
                risk_score += 1
        
        # 违约责任风险
        elif type_code == 'OBJ-08':
            if '违约金上限' not in obligation.content_description:
                risk_score += 1
        
        # 置信度低增加风险
        if obligation.confidence < 50:
            risk_score += 1
            risk_factors.append("条款识别置信度较低")
        
        # 存储风险因素
        obligation.risk_factors = risk_factors
        
        # 转换回风险等级
        if risk_score >= 5:
            return 'critical'
        elif risk_score >= 4:
            return 'high'
        elif risk_score >= 2:
            return 'medium'
        else:
            return 'low'
    
    def detect_product_type(self, clause_text: str, context: Optional[str] = None) -> str:
        """检测产品类型"""
        mapping = self.rules['product_mapping_rules']['general_mapping']['product_type_detection']
        
        full_text = clause_text + (context or "")
        
        for item in mapping:
            for kw in item['keywords']:
                if kw in full_text:
                    return item['type']
        
        return 'unknown'
    
    def compare_with_standard(self, obligation: Obligation, product_type: str) -> Dict[str, Any]:
        """
        与标准产品义务比对
        """
        comparison = {
            'product_type': product_type,
            'differences': []
        }
        
        if product_type not in self.rules['product_mapping_rules']:
            comparison['note'] = '无对应产品类型标准'
            return comparison
        
        product_rules = self.rules['product_mapping_rules'][product_type]
        standards = product_rules.get('standard_obligations', [])
        
        # 简单的差异检测
        obl_type = obligation.obligation_type
        
        # 检查是否有对应的标准义务
        matching_standards = [s for s in standards if s['type'] == obl_type]
        
        if not matching_standards:
            comparison['differences'].append({
                'field': '义务存在性',
                'standard_value': '应有标准义务',
                'contract_value': '合同中存在',
                'impact': 'neutral',
                'note': '非标准义务类型'
            })
            return comparison
        
        standard = matching_standards[0]
        
        # 付款义务特殊比对
        if obl_type == 'OBJ-03':
            comparison = self._compare_payment_obligation(obligation, standard, comparison)
        
        # 维保义务比对
        elif obl_type == 'OBJ-04':
            comparison = self._compare_warranty_obligation(obligation, standard, comparison)
        
        # 交付义务比对
        elif obl_type == 'OBJ-01':
            comparison = self._compare_delivery_obligation(obligation, standard, comparison)
        
        return comparison
    
    def _compare_payment_obligation(self, obligation: Obligation, standard: Dict, comparison: Dict) -> Dict:
        """比对付款义务"""
        content = obligation.content_description
        
        # 预付款比例检查
        if '预付' in content or '预付款' in content:
            match = re.search(r'(\d+(?:\.\d+)?)%', content)
            if match:
                contract_ratio = float(match.group(1))
                standard_ratio = 30  # 标准预付款
                
                if contract_ratio > standard_ratio:
                    comparison['differences'].append({
                        'field': '预付款比例',
                        'standard_value': f'{standard_ratio}%',
                        'contract_value': f'{contract_ratio}%',
                        'impact': 'negative',
                        'note': '预付款比例高于标准，增加资金风险'
                    })
        
        # 质保金检查
        if '质保金' in content:
            match = re.search(r'(\d+(?:\.\d+)?)%', content)
            if match:
                contract_ratio = float(match.group(1))
                standard_ratio = 10
                
                if contract_ratio < standard_ratio:
                    comparison['differences'].append({
                        'field': '质保金比例',
                        'standard_value': f'{standard_ratio}%',
                        'contract_value': f'{contract_ratio}%',
                        'impact': 'negative',
                        'note': '质保金比例低于标准，质量保障不足'
                    })
        
        return comparison
    
    def _compare_warranty_obligation(self, obligation: Obligation, standard: Dict, comparison: Dict) -> Dict:
        """比对维保义务"""
        content = obligation.content_description
        
        # 保修期限检查
        match = re.search(r'(\d+)年', content)
        if match:
            contract_period = int(match.group(1))
            standard_period = 1
            
            if contract_period < standard_period:
                comparison['differences'].append({
                    'field': '保修期限',
                    'standard_value': f'{standard_period}年',
                    'contract_value': f'{contract_period}年',
                    'impact': 'negative',
                    'note': '保修期短于标准'
                })
        
        # 支持等级检查
        if '7×24' in standard.get('standard_clause', ''):
            if '7×24' not in content and '7*24' not in content:
                comparison['differences'].append({
                    'field': '支持等级',
                    'standard_value': '7×24小时',
                    'contract_value': '非7×24小时',
                    'impact': 'negative',
                    'note': '技术支持等级降低'
                })
        
        return comparison
    
    def _compare_delivery_obligation(self, obligation: Obligation, standard: Dict, comparison: Dict) -> Dict:
        """比对交付义务"""
        content = obligation.content_description
        
        # 交付周期检查
        match = re.search(r'(\d+)日', content) or re.search(r'(\d+)天', content)
        if match:
            contract_days = int(match.group(1))
            standard_days = 30
            
            if contract_days > standard_days:
                comparison['differences'].append({
                    'field': '交付周期',
                    'standard_value': f'{standard_days}日内',
                    'contract_value': f'{contract_days}日内',
                    'impact': 'negative',
                    'note': '交付周期长于标准'
                })
        
        return comparison
    
    def process_clause(self, clause_text: str, context: Optional[str] = None) -> List[Obligation]:
        """
        处理单个合同条款，提取履约义务
        """
        obligations = []
        
        # 1. 条款分类
        classifications = self.classify_clause(clause_text)
        
        if not classifications:
            # 无法分类的情况
            return obligations
        
        # 取置信度>15%的所有分类（一条条款可能包含多种义务）
        # 去重+阈值过滤：单条款最多保留Top 2个最高置信度的分类
        MAX_PER_CLAUSE = 2
        MIN_CONFIDENCE = 40  # 提高最小置信度，过滤噪音
        
        # 先按置信度过滤
        filtered_classifications = [
            (ot, conf) for ot, conf in classifications 
            if conf >= MIN_CONFIDENCE
        ]
        
        # 只保留Top N
        filtered_classifications = filtered_classifications[:MAX_PER_CLAUSE]
        
        for obl_type, confidence in filtered_classifications:
            # 低置信度额外检查高权重关键词匹配
            if confidence < 50:
                type_code = obl_type.value
                rules = self.rules['recognition_rules'].get(type_code, {})
                keywords = rules.get('keywords', {})
                has_high_kw = False
                for kw in keywords.get('high_weight', []):
                    if kw in clause_text:
                        has_high_kw = True
                        break
                if not has_high_kw:
                    continue
                
            # 2. 字段提取
            fields = self.extract_fields(clause_text, obl_type)
            
            # 3. 生成义务ID
            type_code = obl_type.value
            self.obligation_counter[type_code] += 1
            obl_id = f"{type_code}-{self.obligation_counter[type_code]:03d}"
            
            # 4. 获取分类信息
            category_info = self.rules['obligation_categories'].get(type_code, {})
            category_name = category_info.get('name', '')
            
            # 5. 提取基本信息
            deadline = self.extract_deadline(clause_text)
            responsible_party = self.extract_responsbile_party(clause_text, type_code)
            
            # 6. 确定子类
            subcategory = self._determine_subcategory(clause_text, type_code)
            
            # 7. 创建义务对象
            obligation = Obligation(
                obligation_id=obl_id,
                obligation_type=type_code,
                category_name=category_name,
                subcategory=subcategory,
                content_description=clause_text.strip()[:200],
                source_text=clause_text.strip(),
                trigger_conditions=self._extract_triggers(clause_text),
                deadline=deadline,
                responsible_party=responsible_party,
                extracted_fields={k: v.value for k, v in fields.items()},
                confidence=confidence,
                notes=""
            )
            
            # 8. 风险评估
            obligation.risk_level = self.determine_risk_level(obligation, fields)
            
            # 9. 标准比对
            product_type = self.detect_product_type(clause_text, context)
            obligation.standard_comparison = self.compare_with_standard(obligation, product_type)
            
            # 10. 风险因素汇总
            obligation.risk_factors = self._collect_risk_factors(obligation)
            
            obligations.append(obligation)
        
        return obligations
    
    def _determine_subcategory(self, clause_text: str, type_code: str) -> str:
        """确定子类"""
        category = self.rules['obligation_categories'].get(type_code, {})
        subcategories = category.get('subcategories', [])
        
        for sub in subcategories:
            if sub[:2] in clause_text or sub in clause_text:
                return sub
        
        return subcategories[0] if subcategories else ""
    
    def _extract_triggers(self, clause_text: str) -> List[str]:
        """提取触发条件"""
        triggers = []
        
        trigger_patterns = [
            r'合同生效后',
            r'验收合格后',
            r'收到发票后',
            r'交付后',
            r'签订后',
            r'满足.*条件后',
        ]
        
        for pattern in trigger_patterns:
            if re.search(pattern, clause_text):
                triggers.append(pattern.replace(r'后', '后').replace(r'.*', ''))
        
        return triggers if triggers else ["合同生效"]
    
    def _collect_risk_factors(self, obligation: Obligation) -> List[str]:
        """收集风险因素"""
        risks = []
        
        # 从标准比对差异中收集
        if obligation.standard_comparison:
            for diff in obligation.standard_comparison.get('differences', []):
                if diff.get('impact') == 'negative':
                    risks.append(diff.get('note', ''))
        
        # 通用风险检查
        if not obligation.deadline:
            risks.append("无明确履行时限")
        
        if '甲方指定地点' in (obligation.content_description or ''):
            risks.append("履行地点不具体")
        
        if obligation.confidence < 50:
            risks.append("条款识别置信度较低")
        
        return risks
    
    def process_contract(self, clauses: List[str], contract_context: Optional[str] = None) -> Dict[str, Any]:
        """
        批量处理合同条款
        """
        all_obligations = []
        
        for i, clause in enumerate(clauses):
            obligations = self.process_clause(clause, contract_context)
            all_obligations.extend(obligations)
        
        # 统计信息
        stats = self._generate_statistics(all_obligations)
        
        return {
            'obligations': [asdict(obl) for obl in all_obligations],
            'statistics': stats,
            'total_count': len(all_obligations)
        }
    
    def _generate_statistics(self, obligations: List[Obligation]) -> Dict[str, Any]:
        """生成统计信息"""
        type_counts = {}
        risk_counts = {}
        
        for obl in obligations:
            # 按类型统计
            type_name = obl.category_name
            type_counts[type_name] = type_counts.get(type_name, 0) + 1
            
            # 按风险统计
            risk_level = obl.risk_level
            risk_counts[risk_level] = risk_counts.get(risk_level, 0) + 1
        
        return {
            'by_type': type_counts,
            'by_risk_level': risk_counts,
            'high_risk_count': risk_counts.get('high', 0) + risk_counts.get('critical', 0)
        }
    
    def export_to_json(self, result: Dict, output_path: str):
        """导出为JSON格式"""
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(result, f, ensure_ascii=False, indent=2)
    
    def export_to_excel(self, result: Dict, output_path: str):
        """导出为Excel格式（需安装openpyxl）"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill
            
            wb = Workbook()
            ws = wb.active
            ws.title = "履约义务清单"
            
            # 获取Excel映射配置
            excel_mapping = self.rules['output_structure']['excel_mapping']
            columns = excel_mapping['columns']
            
            # 写入表头
            for col_config in columns:
                col_letter = col_config['column']
                col_index = ord(col_letter) - ord('A') + 1
                ws.cell(row=1, column=col_index, value=col_config['header'])
            
            # 写入数据
            risk_colors = excel_mapping['risk_highlighting']
            
            for row_idx, obl in enumerate(result['obligations'], start=2):
                for col_config in columns:
                    col_letter = col_config['column']
                    col_index = ord(col_letter) - ord('A') + 1
                    field_name = col_config['field']
                    
                    value = obl.get(field_name, '')
                    cell = ws.cell(row=row_idx, column=col_index, value=value)
                    
                    # 风险等级高亮
                    if field_name == 'risk_level' and value in risk_colors:
                        fill = PatternFill(start_color=risk_colors[value], 
                                          end_color=risk_colors[value], 
                                          fill_type='solid')
                        cell.fill = fill
            
            # 冻结窗格和筛选
            if excel_mapping.get('freeze_panes'):
                ws.freeze_panes = excel_mapping['freeze_panes']
            
            if excel_mapping.get('auto_filter'):
                ws.auto_filter.ref = f"A1:{columns[-1]['column']}{len(result['obligations']) + 1}"
            
            wb.save(output_path)
            return True
        except ImportError:
            print("警告: openpyxl未安装，无法导出Excel")
            return False


def main():
    """测试主函数"""
    print("=" * 60)
    print("履约义务规则库查询接口测试")
    print("=" * 60)
    
    # 初始化查询器
    query = ObligationQuery()
    print(f"✓ 规则库加载成功，版本: {query.rules['metadata']['version']}")
    print(f"✓ 支持义务类型: {len(ObligationType)} 类")
    print()
    
    # 测试条款
    test_clauses = [
        "乙方应在2026年5月30日前将全部设备交付至甲方指定地点，交付时应提供完整的技术文档。",
        "甲方应在设备到货后7日内组织验收，验收标准按照国家相关标准执行，如验收不合格，乙方应在15日内完成整改。",
        "本合同生效后5日内，甲方支付合同总金额的30%作为预付款；验收合格后10日内，支付合同总金额的60%；剩余10%作为质保金，质保期满后支付。",
        "乙方对交付的设备提供3年免费保修服务，保修期内提供7×24小时技术支持，响应时间不超过2小时。",
        "如乙方逾期交付，每逾期一日应按合同总金额的0.05%支付违约金，逾期超过30日的，甲方有权解除合同。",
        "本合同项下开发成果的知识产权归甲方所有，乙方不得向第三方泄露。",
        "乙方应为甲方提供不少于5天的系统操作培训，培训对象为甲方技术人员。",
    ]
    
    print("测试条款:")
    for i, clause in enumerate(test_clauses, 1):
        print(f"{i}. {clause[:60]}...")
    print()
    
    # 处理测试条款
    result = query.process_contract(test_clauses)
    
    print(f"处理结果:")
    print(f"- 提取履约义务总数: {result['total_count']}")
    print(f"- 按类型分布: {result['statistics']['by_type']}")
    print(f"- 按风险分布: {result['statistics']['by_risk_level']}")
    print(f"- 高风险义务数: {result['statistics']['high_risk_count']}")
    print()
    
    # 显示详细结果
    print("详细提取结果:")
    print("-" * 60)
    for obl in result['obligations']:
        print(f"\n【{obl['obligation_id']}】{obl['category_name']} - {obl['subcategory']}")
        print(f"  内容: {obl['content_description'][:80]}...")
        print(f"  责任方: {obl['responsible_party']}")
        print(f"  时限: {obl['deadline'] or '未提取'}")
        print(f"  风险等级: {obl['risk_level']} (置信度: {obl['confidence']:.1f}%)")
        if obl['risk_factors']:
            print(f"  风险因素: {'; '.join(obl['risk_factors'])}")
        if obl['standard_comparison'] and obl['standard_comparison'].get('differences'):
            diff_count = len([d for d in obl['standard_comparison']['differences'] 
                             if d.get('impact') == 'negative'])
            if diff_count > 0:
                print(f"  标准差异: 发现 {diff_count} 项不利差异")
    
    print("\n" + "=" * 60)
    print("测试完成！")
    print("=" * 60)


if __name__ == "__main__":
    main()
