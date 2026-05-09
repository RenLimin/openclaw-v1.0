#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
PDF OCR 合同条款拆分完整流程
直接运行：python3 run_pdf_ocr.py
"""

import sys
import yaml
import re
from pathlib import Path
from typing import List, Dict

# ========== 颜色输出 ==========
class Colors:
    RED = '\033[91m'
    GREEN = '\033[92m'
    YELLOW = '\033[93m'
    CYAN = '\033[96m'
    PURPLE = '\033[95m'
    ENDC = '\033[0m'


# ========== Step 1: PDF OCR ==========
def extract_pdf_text(pdf_path: Path) -> str:
    """智能PDF文本提取：自动检测文本版/扫描版"""
    print(f"\n{Colors.CYAN}{'='*60}{Colors.ENDC}")
    print(f"  📂 解析 PDF: {pdf_path.name}")
    print(f"{Colors.CYAN}{'='*60}{Colors.ENDC}")
    
    # 尝试 PyMuPDF 提取（文本版PDF）
    try:
        import fitz
        doc = fitz.open(str(pdf_path))
        total_chars = 0
        paragraphs = []
        
        for page in doc:
            text = page.get_text()
            total_chars += len(text.strip())
            for para in text.split('\n'):
                para = para.strip()
                if para and len(para) > 2:
                    paragraphs.append(para)
        
        avg_chars = total_chars / max(len(doc), 1)
        if avg_chars > 100:
            print(f"  ✅ 检测为【文本版 PDF】")
            full_text = '\n'.join(paragraphs)
            print(f"  ✅ 提取成功: {len(paragraphs)} 段, {len(full_text)} 字符")
            return full_text
    except Exception as e:
        print(f"  ⚠️  PyMuPDF 提取失败: {e}")
    
    # ========== 扫描版 PDF OCR ==========
    print(f"  📷 检测到扫描版 PDF，启动 OCR ...")
    
    try:
        from pdf2image import convert_from_path
        import pytesseract
    except ImportError as e:
        print(f"  {Colors.RED}❌ 缺少依赖: {e}{Colors.ENDC}")
        print(f"     请执行: pip install pdf2image pytesseract")
        return ""
    
    # PDF 转图片
    print(f"  🖨️  PDF 转图片中 (DPI=300) ...")
    try:
        pages = convert_from_path(str(pdf_path), dpi=300)
        print(f"  ✅ 共 {len(pages)} 页")
    except Exception as e:
        print(f"  {Colors.RED}❌ PDF 转图片失败: {e}{Colors.ENDC}")
        return ""
    
    # OCR 识别
    print(f"  🤖  Tesseract OCR 识别中 (中文+英文) ...")
    all_lines = []
    config = r'--oem 3 --psm 6 -l chi_sim+eng'
    
    for i, page_img in enumerate(pages, 1):
        print(f"  🔍  识别第 {i}/{len(pages)} 页 ...", end='\r')
        text = pytesseract.image_to_string(page_img, config=config)
        for line in text.split('\n'):
            line = line.strip()
            if line and len(line) > 2:
                all_lines.append(line)
    
    full_text = '\n'.join(all_lines)
    print(f"\n  ✅ OCR 完成: {len(all_lines)} 行, {len(full_text)} 字符")
    return full_text


# ========== Step 2: 条款拆分（按边界识别） ==========
def smart_split_clauses(full_text: str) -> List[Dict]:
    print(f"\n{Colors.CYAN}✂️  条款智能拆分中...{Colors.ENDC}")
    
    clause_patterns = [
        r'^\s*第[一二三四五六七八九十\d]+[条章节款]',
        r'^\s*\d+\.\s',
        r'^\s*\d+\.\d+\s',
        r'^\s*[（(][一二三四五六七八九十\d]+[）)]\s*',
        r'^\s*[一二三四五六七八九十]{1,2}[、．.]\s*',
    ]
    combined = re.compile('|'.join(clause_patterns))
    
    lines = full_text.split('\n')
    clauses = []
    current = []
    
    for line in lines:
        line = line.strip()
        if not line or len(line) < 3:
            continue
            
        if combined.match(line) and current:
            content = '\n'.join(current)
            clauses.append({'id': len(clauses)+1, 'content': content, 'length': len(content)})
            current = [line]
        else:
            current.append(line)
    
    if current:
        content = '\n'.join(current)
        clauses.append({'id': len(clauses)+1, 'content': content, 'length': len(content)})
    
    print(f"  ✅ 拆分完成: {len(clauses)} 个条款")
    return clauses


# ========== Step 3: 分类（《民法典》规则） ==========
def classify_clauses(clauses: List[Dict], rules: List[Dict]) -> List[Dict]:
    print(f"\n{Colors.CYAN}🏷️  条款分类中...{Colors.ENDC}")
    
    from typing import Dict as TDict
    category_counts: TDict[str, int] = {}
    
    for idx, clause in enumerate(clauses):
        content = clause['content']
        matched = None
        best_score = -1
        best_prio = 0
        is_early = idx < 4
        
        # ========== 强制匹配规则 ==========
        force_matched = None
        if '[表格' in content or '设备清单' in content or '配置清单' in content:
            for r in rules:
                if r.get('name') == '标的条款': force_matched = r; break
        elif '人民币' in content and '元整' in content and '违约金' not in content:
            for r in rules:
                if r.get('name') == '价格条款': force_matched = r; break
        
        if force_matched:
            clause['category'] = force_matched.get('name', '')
            clause['category_code'] = force_matched.get('code', '')
            category_counts[clause['category']] = category_counts.get(clause['category'], 0) + 1
            continue
        
        # ========== 正常匹配 ==========
        for rule in rules:
            name = rule.get('name', '')
            
            # 排除规则
            exclude = rule.get('exclude_keywords', [])
            should_exclude = False
            for kw in exclude:
                if kw in content: should_exclude = True; break
            if should_exclude: continue
            
            # 计分
            score = 0
            for feat in rule.get('core_features', []):
                if feat in content: score += 3
            for kw in rule.get('keywords', []):
                if kw in content: score += 1
            
            # 位置加权
            if is_early and name == '基本信息': score += 3
            if is_early and name in ['价格条款', '违约责任']: score -= 2
            
            priority = rule.get('priority', 0)
            if (priority > best_prio) or (priority == best_prio and score > best_score):
                best_score = score
                best_prio = priority
                matched = rule
        
        if matched:
            clause['category'] = matched.get('name', '待分类')
            clause['category_code'] = matched.get('code', '99')
        else:
            clause['category'] = '待分类'
            clause['category_code'] = '99'
        
        category_counts[clause['category']] = category_counts.get(clause['category'], 0) + 1
    
    print(f"  {Colors.PURPLE}📊 分类统计:{Colors.ENDC}")
    for cat, cnt in sorted(category_counts.items()):
        print(f"    {cat}: {cnt} 条")
    
    return clauses


# ========== Step 4: 风险识别 ==========
def identify_risks(clauses: List[Dict], risk_rules: Dict) -> List[Dict]:
    print(f"\n{Colors.CYAN}⚠️  风险识别中...{Colors.ENDC}")
    
    high_kws = risk_rules.get('high_risk', {}).get('keywords', [])
    medium_kws = risk_rules.get('medium_risk', {}).get('keywords', [])
    low_kws = risk_rules.get('low_risk', {}).get('keywords', [])
    
    risks = {'high': 0, 'medium': 0, 'low': 0}
    
    for clause in clauses:
        content = clause['content']
        risk_lvl = "✅ 无风险"
        risk_kws = []
        
        for kw in high_kws:
            if kw in content: risk_lvl = "🔴 高风险"; risk_kws.append(kw); break
        if risk_lvl == "✅ 无风险":
            for kw in medium_kws:
                if kw in content: risk_lvl = "🟡 中风险"; risk_kws.append(kw); break
        if risk_lvl == "✅ 无风险":
            for kw in low_kws:
                if kw in content: risk_lvl = "🟢 低风险"; risk_kws.append(kw); break
        
        clause['risk_level'] = risk_lvl
        clause['risk_keywords'] = ', '.join(risk_kws)
        
        if '高风险' in risk_lvl: risks['high'] += 1
        elif '中风险' in risk_lvl: risks['medium'] += 1
        elif '低风险' in risk_lvl: risks['low'] += 1
    
    print(f"  🔴 高风险: {risks['high']} 条")
    print(f"  🟡 中风险: {risks['medium']} 条")
    print(f"  🟢 低风险: {risks['low']} 条")
    return clauses


# ========== Step 5: Excel 导出 ==========
def export_excel(clauses: List[Dict], output: Path, pdf_path: Path):
    print(f"\n{Colors.CYAN}📊 导出 Excel...{Colors.ENDC}")
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print(f"  {Colors.RED}❌ 请安装: pip install openpyxl{Colors.ENDC}")
        return
    
    wb = openpyxl.Workbook()
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    risk_high_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    risk_medium_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    
    # Sheet 1: 全部条款
    ws1 = wb.active
    ws1.title = "全部条款"
    headers = ['条款编号', '条款分类', '风险等级', '风险关键词', '条款内容', '字数']
    for col, h in enumerate(headers, 1):
        c = ws1.cell(row=1, column=col, value=h)
        c.font = header_font; c.fill = header_fill
    
    for row, clause in enumerate(clauses, 2):
        ws1.cell(row=row, column=1, value=clause['id'])
        ws1.cell(row=row, column=2, value=clause['category'])
        ws1.cell(row=row, column=3, value=clause['risk_level'])
        ws1.cell(row=row, column=4, value=clause['risk_keywords'])
        ws1.cell(row=row, column=5, value=clause['content'])
        ws1.cell(row=row, column=6, value=clause['length'])
        
        if '高风险' in clause['risk_level']:
            ws1.cell(row=row, column=3).fill = risk_high_fill
        elif '中风险' in clause['risk_level']:
            ws1.cell(row=row, column=3).fill = risk_medium_fill
    
    ws1.column_dimensions['A'].width = 12
    ws1.column_dimensions['B'].width = 14
    ws1.column_dimensions['C'].width = 12
    ws1.column_dimensions['D'].width = 20
    ws1.column_dimensions['E'].width = 100
    
    # Sheet 2: 分类视图
    ws2 = wb.create_sheet("分类视图")
    cats = sorted(list(set(c['category'] for c in clauses)))
    row = 1
    for cat in cats:
        ws2.cell(row=row, column=1, value=cat)
        ws2.cell(row=row, column=1).font = Font(bold=True, size=14)
        row += 1
        for clause in [c for c in clauses if c['category'] == cat]:
            ws2.cell(row=row, column=1, value=f"条款 {clause['id']}")
            ws2.cell(row=row, column=2, value=clause['risk_level'])
            ws2.cell(row=row, column=3, value=clause['content'][:200])
            row += 1
        row += 1
    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 120
    
    # Sheet 3: 风险汇总
    ws3 = wb.create_sheet("风险条款汇总")
    risk_headers = ['条款编号', '风险等级', '风险关键词', '条款内容']
    for col, h in enumerate(risk_headers, 1):
        c = ws3.cell(row=1, column=col, value=h)
        c.font = header_font; c.fill = header_fill
    
    risk_clauses = [c for c in clauses if '高风险' in c['risk_level'] or '中风险' in c['risk_level']]
    for row, clause in enumerate(risk_clauses, 2):
        ws3.cell(row=row, column=1, value=clause['id'])
        ws3.cell(row=row, column=2, value=clause['risk_level'])
        ws3.cell(row=row, column=3, value=clause['risk_keywords'])
        ws3.cell(row=row, column=4, value=clause['content'])
        if '高风险' in clause['risk_level']:
            ws3.cell(row=row, column=2).fill = risk_high_fill
        elif '中风险' in clause['risk_level']:
            ws3.cell(row=row, column=2).fill = risk_medium_fill
    
    ws3.column_dimensions['A'].width = 12
    ws3.column_dimensions['B'].width = 12
    ws3.column_dimensions['C'].width = 30
    ws3.column_dimensions['D'].width = 120
    
    # Sheet 4: 原始文本
    ws4 = wb.create_sheet("原始文本")
    ws4.cell(row=1, column=1, value="PDF OCR 识别结果（用于核对）")
    ws4.cell(row=1, column=1).font = header_font
    ws4.cell(row=1, column=1).fill = header_fill
    all_text = '\n'.join(c['content'] for c in clauses)
    ws4.cell(row=2, column=1, value=all_text[:32767])
    ws4.column_dimensions['A'].width = 120
    
    wb.save(str(output))
    print(f"  ✅ 已导出: {output}")
    print(f"  📋 工作表: 全部条款, 分类视图, 风险条款汇总, 原始文本")


# ========== 主程序 ==========
def main():
    print(f"\n{Colors.CYAN}╔{'═'*58}╗{Colors.ENDC}")
    print(f"{Colors.CYAN}║   📜 PDF OCR 合同条款拆分工具 v2.0 (Tesseract 引擎)    ║{Colors.ENDC}")
    print(f"{Colors.CYAN}╚{'═'*58}╝{Colors.ENDC}")
    
    pdf_path = Path("/Users/bangcle/Downloads/contract/XSZS2603090130北京聚信得仁.pdf")
    output_path = Path("/Users/bangcle/Downloads/contract/XSZS2603090130_OCR解析结果.xlsx")
    
    # 加载配置
    config_dir = Path("/Users/bangcle/.openclaw/workspace/skills/contract-clause-split/config")
    with open(config_dir / "classification-rules.yaml", 'r', encoding='utf-8') as f:
        class_config = yaml.safe_load(f)
    with open(config_dir / "risk-keywords.yaml", 'r', encoding='utf-8') as f:
        risk_config = yaml.safe_load(f)
    
    rules = class_config.get('classification_rules', [])
    
    # Step 1: PDF OCR
    full_text = extract_pdf_text(pdf_path)
    if not full_text:
        print(f"{Colors.RED}❌ 文本提取失败{Colors.ENDC}")
        return
    
    # Step 2: 条款拆分
    clauses = smart_split_clauses(full_text)
    
    # Step 3: 分类
    clauses = classify_clauses(clauses, rules)
    
    # Step 4: 风险识别
    clauses = identify_risks(clauses, risk_config)
    
    # Step 5: 导出 Excel
    export_excel(clauses, output_path, pdf_path)
    
    print(f"\n{Colors.GREEN}🎉 PDF OCR 解析全部完成！{Colors.ENDC}")
    print(f"   📂 输出文件: {output_path}")
    print(f"\n{Colors.YELLOW}💡 请打开Excel核对OCR识别效果与分类准确性{Colors.ENDC}")


if __name__ == "__main__":
    main()
