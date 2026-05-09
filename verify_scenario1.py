#!/usr/bin/env python3
import openpyxl

# 验证最终版
file_path = 'output/acceptance-test-20260428/scenario1/交付中心项目报告-202602-最终验收通过版.xlsx'
wb = openpyxl.load_workbook(file_path, read_only=True)
print('=== 当前报告实际验证 ===')
print('工作表:', wb.sheetnames[:5])
print('签约表A1值:', repr(wb['签约']['A1'].value))
print('签约表行数:', wb['签约'].max_row)
print()

# 检查几个关键公式是否存在
ws = wb['交付效率统计']
print('交付效率统计B2值:', repr(ws['B2'].value))
print('交付效率统计B3值:', repr(ws['B3'].value))

wb.close()
