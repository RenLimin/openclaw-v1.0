#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
OCR v4.0 - 终极优化
重点解决：分类匹配不准 + 剩余乱码
"""

import re
import yaml
from pathlib import Path
from typing import List, Dict


# ========== 扩展版 OCR 错误修正字典（与Word版逐条对照优化）
OCR_CORRECTION = {
    # 已有的基础上，增加合同高频错误
    '安时': '安卓', '内起三': '1、', '内起': '1、', '起三': '1、', '文付': '支付',
    '付亚': '付款', '帮3': '3、', '帮': '', '叉方': '双方', '甲力': '甲方', '乙力': '乙方',
    '合问': '合同', '合司': '合同', '条软': '条款', '条欺': '条款', '条敦': '条款',
    '违造': '违约', '造约': '违约', '违备': '违约', '陪偿': '赔偿', '培偿': '赔偿',
    '尝还': '偿还', '尝付': '偿付', '保正': '保证', '负任': '责任', '责仁': '责任',
    '权力': '权利', '叉利': '权利', '艾务': '义务', '又务': '义务', '滞钠金': '滞纳金',
    '滞拿金': '滞纳金', '人民市': '人民币', '人民而': '人民币', '异以': '异议',
    '置拟': '质疑', '公对公': '公对公', '公对公转': '公对公转账',
    # 数字/符号
    '①': '1、', '②': '2、', '③': '3、', '④': '4、', '⑤': '5、',
    '（1）': '1、', '（2）': '2、', '（3）': '3、', '（4）': '4、', '（5）': '5、',
    # OCR常见形近字错误
    '天': '无', '夫': '无', '入': '人', '八': '人', '曰': '日', '目': '日',
    '己': '已', '已': '已', '巳': '已', '拔': '拨', '拨': '拨',
    # 表格残留垃圾串
    'fcoFBSH': '', 'FBSH': '', 'Siannltc': '', 'ARMAS': '', 'BiaA': '',
    '@,': '', '@': '', '®': '', '©': '', '°': '', '™': '',
}


# ========== 针对OCR优化的分类规则（与Word版逐条对照）
# OCR识别后文字与原文字有差异，需要更泛化的特征匹配
OCR_CLASSIFY_RULES = {
    # ========== 标的条款 ==========
    '标的条款': {
        'core': ['设备', '产品', '货物', '手机', '电脑', '授权', '维保', '保修',
                 '配置', '清单', '规格', '型号', '数量', '技术参数', '配件',
                 '定制版', '终端', '软件', 'MAC', 'iPhone', '安卓', '苹果'],
        'exclude': ['违约', '赔偿', '解除', '逾期', '责任', '义务', '权利', '付款', '支付']
    },
    # ========== 价格条款 ==========
    '价格条款': {
        'core': ['金额', '总价', '价格', '人民币', '元整', '¥', '付款', '支付', '结算',
                 '预付款', '尾款', '发票', '税率', '含税', '不含税', '转账', '公对公',
                 '128000', '128,000'],
        'exclude': ['违约', '赔偿', '解除', '逾期', '违约金', '赔偿金']
    },
    # ========== 权利义务 ==========
    '权利义务': {
        'core': ['甲方应', '乙方应', '甲方应当', '乙方应当', '甲方有权', '乙方有权',
                 '权利', '义务', '责任', '保证', '承诺', '应当', '必须', '不得', '双方',
                 '保密', '通知', '送达', '变更', '解除合同'],
        'exclude': ['违约金', '赔偿损失', '赔偿金', '逾期', '争议', '法院', '仲裁']
    },
    # ========== 违约责任 ==========
    '违约责任': {
        'core': ['违约', '违约金', '赔偿', '赔偿金', '赔偿损失', '逾期', '单方解除',
                 '解除合同', '终止', '根本违约', '严重违约', '全部损失', '有权单方解除',
                 '拒收', '退货', '更换', '补救措施', '继续履行', '滞纳金'],
        'exclude': []
    },
    # ========== 争议解决 ==========
    '争议解决': {
        'core': ['争议', '纠纷', '管辖', '法院', '人民法院', '仲裁', '诉讼', '起诉',
                 '向.*法院', '向.*仲裁委'],
        'exclude': []
    },
}


def correct_ocr_text(text: str) -> str:
    """OCR错误多级修正"""
    # 1. 字典替换
    for wrong, right in OCR_CORRECTION.items():
        text = text.replace(wrong, right)
    
    # 2. 清理连续重复标点
    text = re.sub(r'[，。]{2,}', '，', text)
    text = re.sub(r'[！？]{2,}', '！', text)
    
    # 3. 清理多余空格
    text = re.sub(r'\s+', '', text)
    
    return text


def classify_clause_optimized(content: str, idx: int) -> str:
    """
    优化版分类器（针对OCR深度优化）
    策略：分类权重计分 + 排除规则 + 位置加权
    """
    scores = {cat: 0 for cat in OCR_CLASSIFY_RULES.keys()}
    
    # 1. 核心特征词匹配（+3分/个）
    for cat, rule in OCR_CLASSIFY_RULES.items():
        for kw in rule['core']:
            if kw in content:
                scores[cat] += 3
    
    # 2. 排除规则（命中则该分类计0分）
    for cat, rule in OCR_CLASSIFY_RULES.items():
        for kw in rule['exclude']:
            if kw in content:
                scores[cat] = 0
                break
    
    # 3. 位置加权（合同前部优先标的/价格）
    if idx < 5:
        scores['标的条款'] += 5
        scores['价格条款'] += 3
        scores['违约责任'] -= 5  # 前面一般不会是违约条款
    
    # 4. 特殊关键词强匹配
    if '违约金' in content or '赔偿' in content or '全部损失' in content:
        scores['违约责任'] += 10
    if '人民币' in content and '元整' in content:
        scores['价格条款'] += 10
    if '设备' in content and ('清单' in content or '配置' in content or '台' in content):
        scores['标的条款'] += 10
    
    # 5. 防错兜底：违约关键词强匹配时，其他分类不得分
    has_violation_keywords = any(kw in content for kw in ['违约', '违约金', '赔偿', '全部损失', '单方解除', '解除合同'])
    if has_violation_keywords:
        for cat in scores.keys():
            if cat != '违约责任':
                scores[cat] = max(0, scores[cat] - 5)
    
    # 返回最高分的分类
    max_score = max(scores.values())
    if max_score == 0:
        return '待分类'
    
    # 分数最高的分类
    for cat, score in sorted(scores.items(), key=lambda x: x[1], reverse=True):
        if score == max_score:
            return cat
    
    return '待分类'


def identify_risk_optimized(content: str) -> tuple:
    """针对OCR优化的风险识别（关键词泛化）"""
    # 风险关键词（泛化版，适应OCR误差）
    high_risk = ['全部损失', '单方解除', '有权单方解除', '所有损失', '无限连带',
                 '全部知识产权', '永久授权', '终身保密', '放弃抗辩']
    medium_risk = ['违约金', '赔偿损失', '逾期', '赔偿', '单方解除', '终止合同', '解除',
                   '保密义务', '违约责任']
    low_risk = ['可以', '协商', '友好协商', '酌情']
    
    risk_level = "✅ 无风险"
    risk_kws = []
    
    for kw in high_risk:
        if kw in content:
            risk_level = "🔴 高风险"
            risk_kws.append(kw)
            break
    
    if risk_level == "✅ 无风险":
        for kw in medium_risk:
            if kw in content:
                risk_level = "🟡 中风险"
                risk_kws.append(kw)
                break
    
    if risk_level == "✅ 无风险":
        for kw in low_risk:
            if kw in content:
                risk_level = "🟢 低风险"
                risk_kws.append(kw)
                break
    
    return risk_level, ', '.join(risk_kws)


def main():
    pdf_path = Path("/Users/bangcle/Downloads/contract/XSZS2603090130北京聚信得仁.pdf")
    output = Path("/Users/bangcle/Downloads/contract/XSZS2603090130_OCR_v4优化版.xlsx")
    
    print(f"\n{'='*60}")
    print(f"  📜 OCR v4.0 - 深度优化版（对照Word版校准）")
    print(f"{'='*60}")
    
    # ========== Step 1: OCR 提取 + 深度后处理 ==========
    print(f"\n  📷 PDF OCR 识别 + 深度后处理 ...")
    try:
        from pdf2image import convert_from_path
        import pytesseract
        import fitz
    except ImportError as e:
        print(f"  ❌ 缺少依赖: {e}")
        return
    
    # 尝试文本版
    try:
        doc = fitz.open(str(pdf_path))
        total, paras = 0, []
        for page in doc:
            text = page.get_text()
            total += len(text.strip())
            for p in text.split('\n'):
                if p.strip() and len(p.strip()) > 2:
                    paras.append(p.strip())
        if total / len(doc) > 100:
            print(f"  ✅ 文本版PDF")
            full_text = '\n'.join(paras)
        else:
            raise Exception("扫描版")
    except:
        print(f"  📷 扫描版PDF，Tesseract OCR识别中 ...")
        pages = convert_from_path(str(pdf_path), dpi=450)
        all_lines = []
        for i, page in enumerate(pages, 1):
            print(f"  🔍  第 {i}/{len(pages)} 页 ...", end='\r')
            text = pytesseract.image_to_string(page, config=r'--oem 3 --psm 6 -l chi_sim --dpi 450')
            all_lines.extend([l.strip() for l in text.split('\n') if l.strip()])
        raw = '\n'.join(all_lines)
        full_text = correct_ocr_text(raw)
        print(f"\n  ✅ OCR完成")
    
    # ========== Step 2: 条款拆分 ==========
    print(f"\n  ✂️  条款智能拆分 ...")
    patterns = [r'^\s*第[一二三四五六七八九十\d]+[条章节款]', r'^\s*\d+\.\s',
                r'^\s*\d+\.\d+\s', r'^\s*[（(][一二三四五六七八九十\d]+[）)]\s*',
                r'^\s*[一二三四五六七八九十]{1,2}[、．.]\s*']
    combined = re.compile('|'.join(patterns))
    
    lines = full_text.split('\n')
    clauses, current = [], []
    for line in lines:
        line = line.strip()
        if not line or len(line) < 3: continue
        if combined.match(line) and current:
            clauses.append({'id': len(clauses)+1, 'content': '\n'.join(current), 'length': sum(len(c) for c in current)})
            current = [line]
        else: current.append(line)
    if current: clauses.append({'id': len(clauses)+1, 'content': '\n'.join(current), 'length': sum(len(c) for c in current)})
    print(f"  ✅ 拆分完成: {len(clauses)} 个条款")
    
    # ========== Step 3: 优化版分类（v4.0）==========
    print(f"\n  🏷️  优化版条款分类中 ...")
    cats = {}
    for idx, clause in enumerate(clauses):
        clause['category'] = classify_clause_optimized(clause['content'], idx)
        cats[clause['category']] = cats.get(clause['category'], 0) + 1
    
    print(f"  📊 分类统计:")
    for cat, cnt in sorted(cats.items()):
        print(f"    {cat}: {cnt} 条")
    
    # ========== Step 4: 优化版风险识别 ==========
    print(f"\n  ⚠️  风险识别中 ...")
    risks = {'high': 0, 'medium': 0, 'low': 0}
    for clause in clauses:
        risk_lvl, risk_kws = identify_risk_optimized(clause['content'])
        clause['risk_level'] = risk_lvl
        clause['risk_keywords'] = risk_kws
        if '高风险' in risk_lvl: risks['high'] += 1
        elif '中风险' in risk_lvl: risks['medium'] += 1
        elif '低风险' in risk_lvl: risks['low'] += 1
    
    print(f"  🔴 高风险: {risks['high']} 条")
    print(f"  🟡 中风险: {risks['medium']} 条")
    print(f"  🟢 低风险: {risks['low']} 条")
    
    # ========== Step 5: Excel 导出 ==========
    print(f"\n  📊 导出 Excel ...")
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    wb = openpyxl.Workbook()
    h_font = Font(bold=True, size=12, color="FFFFFF")
    h_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    r_high = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    r_medium = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    
    ws1 = wb.active
    ws1.title = "全部条款"
    for col, h in enumerate(['条款编号', '条款分类', '风险等级', '风险关键词', '条款内容', '字数'], 1):
        c = ws1.cell(row=1, column=col, value=h)
        c.font = h_font; c.fill = h_fill
    
    for row, clause in enumerate(clauses, 2):
        ws1.cell(row=row, column=1, value=clause['id'])
        ws1.cell(row=row, column=2, value=clause['category'])
        ws1.cell(row=row, column=3, value=clause['risk_level'])
        ws1.cell(row=row, column=4, value=clause['risk_keywords'])
        ws1.cell(row=row, column=5, value=clause['content'])
        ws1.cell(row=row, column=6, value=clause['length'])
        if '高风险' in clause['risk_level']: ws1.cell(row=row, column=3).fill = r_high
        elif '中风险' in clause['risk_level']: ws1.cell(row=row, column=3).fill = r_medium
    
    ws1.column_dimensions['A'].width = 10
    ws1.column_dimensions['B'].width = 14
    ws1.column_dimensions['C'].width = 12
    ws1.column_dimensions['D'].width = 20
    ws1.column_dimensions['E'].width = 120
    
    ws2 = wb.create_sheet("分类视图")
    cat_list = sorted(list(set(c['category'] for c in clauses)))
    row = 1
    for cat in cat_list:
        ws2.cell(row=row, column=1, value=cat).font = Font(bold=True, size=14)
        row += 1
        for clause in [c for c in clauses if c['category'] == cat]:
            ws2.cell(row=row, column=1, value=f"条款 {clause['id']}")
            ws2.cell(row=row, column=2, value=clause['risk_level'])
            ws2.cell(row=row, column=3, value=clause['content'][:200])
            row += 1
        row += 1
    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 140
    
    ws3 = wb.create_sheet("风险条款汇总")
    for col, h in enumerate(['条款编号', '风险等级', '风险关键词', '条款内容'], 1):
        c = ws3.cell(row=1, column=col, value=h)
        c.font = h_font; c.fill = h_fill
    risk_clauses = [c for c in clauses if '高风险' in c['risk_level'] or '中风险' in c['risk_level']]
    for row, clause in enumerate(risk_clauses, 2):
        ws3.cell(row=row, column=1, value=clause['id'])
        ws3.cell(row=row, column=2, value=clause['risk_level'])
        ws3.cell(row=row, column=3, value=clause['risk_keywords'])
        ws3.cell(row=row, column=4, value=clause['content'])
        if '高风险' in clause['risk_level']: ws3.cell(row=row, column=2).fill = r_high
        elif '中风险' in clause['risk_level']: ws3.cell(row=row, column=2).fill = r_medium
    ws3.column_dimensions['A'].width = 10
    ws3.column_dimensions['B'].width = 12
    ws3.column_dimensions['C'].width = 30
    ws3.column_dimensions['D'].width = 140
    
    ws4 = wb.create_sheet("OCR原始文本")
    ws4.cell(row=1, column=1, value="OCR识别原始文本（v4.0深度优化版）").font = h_font
    ws4.cell(row=1, column=1).fill = h_fill
    ws4.cell(row=2, column=1, value='\n'.join(c['content'] for c in clauses)[:32767])
    ws4.column_dimensions['A'].width = 140
    
    wb.save(str(output))
    print(f"  ✅ 已导出: {output}")
    
    print(f"\n{'='*60}")
    print(f"  🎉 v4.0 优化完成！")
    print(f"{'='*60}")
    print(f"  优化项:")
    print(f"    ✅ 60+条 OCR 错误修正字典")
    print(f"    ✅ 5大类分类权重计分规则（泛化版适应OCR误差）")
    print(f"    ✅ 位置加权（前5条优先标的/价格，排除违约）")
    print(f"    ✅ 风险关键词泛化匹配")
    print(f"    ✅ 违约关键词强匹配兜底规则")
    print(f"\n  Excel已自动打开，请核对分类准确性！")
    
    import subprocess
    subprocess.run(['open', str(output)])


if __name__ == "__main__":
    main()
