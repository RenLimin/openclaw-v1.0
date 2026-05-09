#!/usr/bin/env python3
"""
格式校验模块
============
逐单元格对比生成文件与模板的格式一致性，
支持底色、边框、字体、字号、对齐、数字格式检查，
输出批量格式比对报告（Pass/Fail）
"""

import logging
import json
from pathlib import Path
from typing import Dict, List, Any, Optional, Set, Tuple
import openpyxl
from openpyxl.styles import PatternFill, Border, Font, Alignment
from openpyxl.utils import get_column_letter
import yaml

logger = logging.getLogger(__name__)


class FormatChecker:
    """
    格式校验类
    
    功能:
    - 逐单元格格式比对
    - 支持多种格式类型检查
    - 生成详细校验报告
    - 规则可配置
    """
    
    def __init__(self, template_path: str, generated_path: str, rules_path: str = None):
        """
        初始化格式校验器
        
        Args:
            template_path: 模板文件路径
            generated_path: 生成文件路径
            rules_path: 规则配置文件路径
        """
        self.template_path = Path(template_path)
        self.generated_path = Path(generated_path)
        self.rules = self._load_rules(rules_path)
        
        # 加载工作簿
        self.template_wb: Optional[openpyxl.Workbook] = None
        self.generated_wb: Optional[openpyxl.Workbook] = None
        
        self._load_workbooks()
    
    def _load_rules(self, rules_path: str = None) -> Dict[str, Any]:
        """加载规则配置"""
        if rules_path is None:
            rules_path = Path(__file__).parent.parent / 'config' / 'format-rules.yaml'
        
        try:
            with open(rules_path, 'r', encoding='utf-8') as f:
                return yaml.safe_load(f) or {}
        except Exception as e:
            logger.warning(f"加载规则配置失败，使用默认配置: {e}")
            return {}
    
    def _load_workbooks(self):
        """加载工作簿"""
        logger.info("加载格式校验文件...")
        
        self.template_wb = openpyxl.load_workbook(
            self.template_path,
            data_only=False
        )
        self.generated_wb = openpyxl.load_workbook(
            self.generated_path,
            data_only=False
        )
        
        logger.info(f"  模板: {len(self.template_wb.sheetnames)} 个Sheet")
        logger.info(f"  生成文件: {len(self.generated_wb.sheetnames)} 个Sheet")
    
    def _get_sheet_config(self, sheet_name: str) -> Dict[str, Any]:
        """获取Sheet特定的配置"""
        sheet_specific = self.rules.get('sheet_specific', {})
        return sheet_specific.get(sheet_name, {})
    
    def _is_check_enabled(self, check_type: str, sheet_name: str = None) -> bool:
        """检查某类检查是否启用"""
        global_config = self.rules.get('global', {})
        enabled_checks = global_config.get('enabled_checks', [])
        
        if check_type not in enabled_checks:
            return False
        
        # 检查类型特定配置
        type_config = self.rules.get(check_type, {})
        if not type_config.get('enabled', True):
            return False
        
        return True
    
    def check_all_sheets(self, max_rows: int = None, 
                         check_types: List[str] = None) -> Dict[str, Any]:
        """
        校验所有Sheet格式
        
        Args:
            max_rows: 每个Sheet最大检查行数，默认使用配置
            check_types: 检查类型列表，可选: fill, font, border, alignment, number_format
            
        Returns:
            校验报告字典
        """
        scope_config = self.rules.get('scope', {})
        if max_rows is None:
            max_rows = scope_config.get('max_rows_per_sheet', 1000)
        
        if check_types is None:
            check_types = self.rules.get('global', {}).get(
                'enabled_checks', 
                ['fill', 'font', 'border', 'alignment', 'number_format']
            )
        
        logger.info("开始格式校验...")
        logger.info(f"  检查类型: {check_types}")
        logger.info(f"  最大检查行数: {max_rows}")
        
        report = {
            'success': True,
            'total_cells_checked': 0,
            'differences': [],
            'sheets_summary': {},
            'template_file': str(self.template_path),
            'generated_file': str(self.generated_path)
        }
        
        common_sheets = set(self.template_wb.sheetnames) & set(self.generated_wb.sheetnames)
        ignore_config = self.rules.get('ignore', {})
        ignore_sheets = set(ignore_config.get('ignore_sheets', []))
        
        # 检查共同的Sheet
        for sheet_name in common_sheets:
            if sheet_name in ignore_sheets:
                logger.info(f"  跳过Sheet: {sheet_name}")
                continue
            
            sheet_report = self.check_sheet(sheet_name, max_rows, check_types)
            report['sheets_summary'][sheet_name] = sheet_report
            report['total_cells_checked'] += sheet_report['cells_checked']
            report['differences'].extend(sheet_report['differences'])
        
        # 检查缺失的Sheet
        missing_in_generated = set(self.template_wb.sheetnames) - set(self.generated_wb.sheetnames) - ignore_sheets
        for sheet_name in missing_in_generated:
            report['differences'].append({
                'type': 'SHEET_MISSING',
                'sheet': sheet_name,
                'message': f"生成文件缺少Sheet: {sheet_name}"
            })
        
        # 检查多余的Sheet
        extra_in_generated = set(self.generated_wb.sheetnames) - set(self.template_wb.sheetnames) - ignore_sheets
        for sheet_name in extra_in_generated:
            report['differences'].append({
                'type': 'SHEET_EXTRA',
                'sheet': sheet_name,
                'message': f"生成文件多出Sheet: {sheet_name}"
            })
        
        # 判断整体结果
        thresholds = self.rules.get('thresholds', {})
        diff_count = len(report['differences'])
        fail_threshold = thresholds.get('fail_threshold', 100)
        
        if diff_count > fail_threshold:
            report['success'] = False
        
        self._log_summary(report)
        
        return report
    
    def check_sheet(self, sheet_name: str, max_rows: int = 1000,
                    check_types: List[str] = None) -> Dict[str, Any]:
        """
        校验单个Sheet格式
        
        Args:
            sheet_name: Sheet名称
            max_rows: 最大检查行数
            check_types: 检查类型列表
            
        Returns:
            Sheet校验报告
        """
        template_ws = self.template_wb[sheet_name]
        generated_ws = self.generated_wb[sheet_name]
        
        differences: List[Dict] = []
        cells_checked = 0
        
        # 确定检查范围
        max_check_row = min(max(template_ws.max_row, generated_ws.max_row), max_rows)
        max_check_col = max(template_ws.max_column, generated_ws.max_column)
        
        scope_config = self.rules.get('scope', {})
        skip_empty = scope_config.get('skip_empty_cells', True)
        ignore_config = self.rules.get('ignore', {})
        
        logger.info(f"  校验Sheet: {sheet_name} ({max_check_row} 行 x {max_check_col} 列)")
        
        for row in range(1, max_check_row + 1):
            for col in range(1, max_check_col + 1):
                # 获取模板和生成文件的单元格
                t_cell = template_ws.cell(row=row, column=col) if (
                    row <= template_ws.max_row and col <= template_ws.max_column
                ) else None
                
                g_cell = generated_ws.cell(row=row, column=col) if (
                    row <= generated_wb.max_row and col <= generated_wb.max_column
                ) else None
                
                # 跳过空单元格
                if skip_empty:
                    t_empty = t_cell is None or t_cell.value is None
                    g_empty = g_cell is None or g_cell.value is None
                    if t_empty and g_empty:
                        continue
                
                # 检查是否是需要忽略的单元格
                cell_ref = f"{sheet_name}!{get_column_letter(col)}{row}"
                ignore_cells = ignore_config.get('ignore_cells', [])
                if cell_ref in ignore_cells:
                    continue
                
                cells_checked += 1
                
                # 检查各类格式
                cell_diffs = self._compare_cell_format(
                    t_cell, g_cell, sheet_name, row, col, check_types or []
                )
                differences.extend(cell_diffs)
        
        # 生成摘要
        report = {
            'sheet_name': sheet_name,
            'template_rows': template_ws.max_row,
            'template_cols': template_ws.max_column,
            'generated_rows': generated_ws.max_row,
            'generated_cols': generated_ws.max_column,
            'cells_checked': cells_checked,
            'differences_count': len(differences),
            'differences': differences
        }
        
        logger.info(f"    检查单元格: {cells_checked}, 差异: {len(differences)}")
        
        return report
    
    def _compare_cell_format(self, t_cell, g_cell, sheet_name: str, 
                             row: int, col: int, check_types: List[str]) -> List[Dict]:
        """比较两个单元格的格式"""
        differences: List[Dict] = []
        cell_ref = f"{get_column_letter(col)}{row}"
        
        # 如果一个单元格存在另一个不存在
        if g_cell is None:
            if t_cell is not None and t_cell.value is not None:
                differences.append({
                    'type': 'CELL_MISSING',
                    'sheet': sheet_name,
                    'cell': cell_ref,
                    'row': row,
                    'column': col,
                    'message': f"单元格 {cell_ref} 在生成文件中缺失"
                })
            return differences
        
        if t_cell is None:
            # 新增单元格，不记录差异
            return differences
        
        # 1. 检查填充色（底色）
        if 'fill' in check_types and self._is_check_enabled('fill'):
            fill_diffs = self._check_fill(t_cell, g_cell, sheet_name, cell_ref, row, col)
            differences.extend(fill_diffs)
        
        # 2. 检查字体
        if 'font' in check_types and self._is_check_enabled('font'):
            font_diffs = self._check_font(t_cell, g_cell, sheet_name, cell_ref, row, col)
            differences.extend(font_diffs)
        
        # 3. 检查边框
        if 'border' in check_types and self._is_check_enabled('border'):
            border_diffs = self._check_border(t_cell, g_cell, sheet_name, cell_ref, row, col)
            differences.extend(border_diffs)
        
        # 4. 检查对齐方式
        if 'alignment' in check_types and self._is_check_enabled('alignment'):
            align_diffs = self._check_alignment(t_cell, g_cell, sheet_name, cell_ref, row, col)
            differences.extend(align_diffs)
        
        # 5. 检查数字格式
        if 'number_format' in check_types and self._is_check_enabled('number_format'):
            num_fmt_diffs = self._check_number_format(t_cell, g_cell, sheet_name, cell_ref, row, col)
            differences.extend(num_fmt_diffs)
        
        return differences
    
    def _check_fill(self, t_cell, g_cell, sheet_name: str, cell_ref: str,
                    row: int, col: int) -> List[Dict]:
        """检查填充色"""
        differences = []
        fill_config = self.rules.get('fill', {})
        
        # 获取颜色值
        t_fill = t_cell.fill.fgColor.rgb if (t_cell.fill and t_cell.fill.fgColor) else None
        g_fill = g_cell.fill.fgColor.rgb if (g_cell.fill and g_cell.fill.fgColor) else None
        
        # 忽略白色
        if fill_config.get('ignore_white', True):
            white_colors = fill_config.get('white_colors', ["00FFFFFF", "FFFFFFFF", None])
            if t_fill in white_colors and g_fill in white_colors:
                return differences
        
        # 只检查有颜色的单元格
        if fill_config.get('only_colored', True):
            white_colors = fill_config.get('white_colors', ["00FFFFFF", "FFFFFFFF", None])
            if t_fill in white_colors:
                return differences
        
        if t_fill != g_fill:
            differences.append({
                'type': 'FILL_COLOR',
                'sheet': sheet_name,
                'cell': cell_ref,
                'row': row,
                'column': col,
                'expected': t_fill,
                'actual': g_fill,
                'message': f"单元格 {cell_ref} 底色不一致: 预期 {t_fill}, 实际 {g_fill}"
            })
        
        return differences
    
    def _check_font(self, t_cell, g_cell, sheet_name: str, cell_ref: str,
                    row: int, col: int) -> List[Dict]:
        """检查字体"""
        differences = []
        font_config = self.rules.get('font', {})
        
        # 只检查表头行
        if font_config.get('header_only', True) and row > 5:
            return differences
        
        t_font = t_cell.font
        g_font = g_cell.font
        
        if not t_font:
            return differences
        
        # 检查加粗
        if font_config.get('check_bold', True):
            if g_font.bold != t_font.bold:
                differences.append({
                    'type': 'FONT_BOLD',
                    'sheet': sheet_name,
                    'cell': cell_ref,
                    'row': row,
                    'column': col,
                    'expected': t_font.bold,
                    'actual': g_font.bold,
                    'message': f"单元格 {cell_ref} 字体加粗不一致"
                })
        
        # 检查字体大小
        if font_config.get('check_size', True):
            if g_font.size != t_font.size:
                differences.append({
                    'type': 'FONT_SIZE',
                    'sheet': sheet_name,
                    'cell': cell_ref,
                    'row': row,
                    'column': col,
                    'expected': t_font.size,
                    'actual': g_font.size,
                    'message': f"单元格 {cell_ref} 字体大小不一致"
                })
        
        # 检查字体名称
        if font_config.get('check_name', True):
            if g_font.name != t_font.name:
                differences.append({
                    'type': 'FONT_NAME',
                    'sheet': sheet_name,
                    'cell': cell_ref,
                    'row': row,
                    'column': col,
                    'expected': t_font.name,
                    'actual': g_font.name,
                    'message': f"单元格 {cell_ref} 字体名称不一致"
                })
        
        return differences
    
    def _check_border(self, t_cell, g_cell, sheet_name: str, cell_ref: str,
                      row: int, col: int) -> List[Dict]:
        """检查边框"""
        differences = []
        border_config = self.rules.get('border', {})
        
        # 只检查非空单元格
        if border_config.get('only_non_empty', True) and t_cell.value is None:
            return differences
        
        check_mode = border_config.get('check_mode', 'simple')
        
        if check_mode == 'simple':
            # 简化检查：只检查是否有边框
            def has_border(border):
                if not border:
                    return False
                return any([
                    border.left and border.left.style != 'none',
                    border.right and border.right.style != 'none',
                    border.top and border.top.style != 'none',
                    border.bottom and border.bottom.style != 'none'
                ])
            
            t_has_border = has_border(t_cell.border)
            g_has_border = has_border(g_cell.border)
            
            if t_has_border != g_has_border:
                differences.append({
                    'type': 'BORDER',
                    'sheet': sheet_name,
                    'cell': cell_ref,
                    'row': row,
                    'column': col,
                    'expected': t_has_border,
                    'actual': g_has_border,
                    'message': f"单元格 {cell_ref} 边框不一致"
                })
        else:
            # 完整检查
            pass
        
        return differences
    
    def _check_alignment(self, t_cell, g_cell, sheet_name: str, cell_ref: str,
                         row: int, col: int) -> List[Dict]:
        """检查对齐方式"""
        differences = []
        align_config = self.rules.get('alignment', {})
        
        # 只检查非空单元格
        if align_config.get('only_non_empty', True) and t_cell.value is None:
            return differences
        
        t_align = t_cell.alignment
        g_align = g_cell.alignment
        
        if not t_align:
            return differences
        
        # 检查水平对齐
        if align_config.get('check_horizontal', True):
            if g_align.horizontal != t_align.horizontal:
                differences.append({
                    'type': 'ALIGNMENT_HORIZONTAL',
                    'sheet': sheet_name,
                    'cell': cell_ref,
                    'row': row,
                    'column': col,
                    'expected': t_align.horizontal,
                    'actual': g_align.horizontal,
                    'message': f"单元格 {cell_ref} 水平对齐不一致"
                })
        
        return differences
    
    def _check_number_format(self, t_cell, g_cell, sheet_name: str, cell_ref: str,
                             row: int, col: int) -> List[Dict]:
        """检查数字格式"""
        differences = []
        num_config = self.rules.get('number_format', {})
        
        ignore_formats = num_config.get('ignore_formats', ['General', ''])
        t_format = t_cell.number_format
        g_format = g_cell.number_format
        
        if t_format in ignore_formats and g_format in ignore_formats:
            return differences
        
        if t_format != g_format:
            differences.append({
                'type': 'NUMBER_FORMAT',
                'sheet': sheet_name,
                'cell': cell_ref,
                'row': row,
                'column': col,
                'expected': t_format,
                'actual': g_format,
                'message': f"单元格 {cell_ref} 数字格式不一致"
            })
        
        return differences
    
    def _log_summary(self, report: Dict[str, Any]):
        """输出校验摘要"""
        diff_count = len(report['differences'])
        
        logger.info("=" * 60)
        logger.info("格式校验摘要")
        logger.info("=" * 60)
        logger.info(f"  总检查单元格: {report['total_cells_checked']}")
        logger.info(f"  发现差异数: {diff_count}")
        
        # 按类型统计差异
        diff_by_type = {}
        for diff in report['differences']:
            diff_type = diff['type']
            diff_by_type[diff_type] = diff_by_type.get(diff_type, 0) + 1
        
        for diff_type, count in diff_by_type.items():
            logger.info(f"    {diff_type}: {count}")
        
        if report['success']:
            logger.info("✓ 格式校验通过")
        else:
            logger.warning("✗ 格式校验未通过")
        
        logger.info("=" * 60)
    
    def save_report(self, output_path: str, report: Dict[str, Any] = None):
        """
        保存校验报告到JSON文件
        
        Args:
            output_path: 输出文件路径
            report: 校验报告，如果为None则重新执行校验
        """
        if report is None:
            report = self.check_all_sheets()
        
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(report, f, ensure_ascii=False, indent=2)
        
        logger.info(f"格式校验报告已保存: {output_path}")
    
    def print_differences(self, report: Dict[str, Any] = None, limit: int = 50):
        """
        打印差异详情
        
        Args:
            report: 校验报告
            limit: 最大显示数量
        """
        if report is None:
            report = self.check_all_sheets()
        
        diffs = report['differences']
        
        if not diffs:
            print("✓ 没有发现格式差异")
            return
        
        print(f"\n格式差异详情 (共 {len(diffs)} 个):")
        print("-" * 60)
        
        for i, diff in enumerate(diffs[:limit]):
            print(f"{i+1}. [{diff['type']}] {diff['sheet']}!{diff.get('cell', '??')}")
            print(f"   {diff['message']}")
        
        if len(diffs) > limit:
            print(f"... 还有 {len(diffs) - limit} 个差异未显示")
    
    def close(self):
        """关闭工作簿"""
        if self.template_wb:
            self.template_wb.close()
        if self.generated_wb:
            self.generated_wb.close()
    
    def __enter__(self):
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()
