#!/usr/bin/env python3
"""
Excel IO 高性能读写模块
======================
支持 10万+ 行级数据高效读写（openpyxl / pandas 双后端）
CSV/Excel 自动格式检测，多 Sheet 批量读写，数据类型自动识别与转换
"""

import logging
from pathlib import Path
from typing import Dict, List, Any, Optional, Union
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils import get_column_letter
import yaml

logger = logging.getLogger(__name__)


class ExcelIO:
    """
    Excel/CSV 高性能读写类
    
    功能:
    - 自动检测文件格式
    - 大文件优化读取
    - 多 Sheet 批量操作
    - 数据类型自动转换
    """
    
    def __init__(self, config_path: str = None):
        """
        初始化 Excel IO
        
        Args:
            config_path: 配置文件路径，默认使用 ../config/data-import.yaml
        """
        self.config = self._load_config(config_path)
        self._large_file_threshold = self.config.get(
            'performance', {}
        ).get('large_file_threshold', 10000)
    
    def _load_config(self, config_path: str = None) -> Dict[str, Any]:
        """加载配置文件"""
        if config_path is None:
            config_path = Path(__file__).parent.parent / 'config' / 'data-import.yaml'
        
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                return yaml.safe_load(f) or {}
        except Exception as e:
            logger.warning(f"加载配置文件失败，使用默认配置: {e}")
            return {}
    
    def detect_format(self, file_path: str) -> str:
        """
        检测文件格式
        
        Args:
            file_path: 文件路径
            
        Returns:
            'excel' 或 'csv'
        """
        ext = Path(file_path).suffix.lower()
        
        if ext in ['.xlsx', '.xlsm', '.xls', '.xltx', '.xltm']:
            return 'excel'
        elif ext in ['.csv', '.tsv', '.txt']:
            return 'csv'
        else:
            # 尝试读取内容检测
            try:
                with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                    first_line = f.readline(1000)
                    if ',' in first_line or '\t' in first_line:
                        return 'csv'
            except:
                pass
            return 'excel'  # 默认按Excel尝试
    
    def read_file(self, file_path: str, sheet_name: str = None, **kwargs) -> pd.DataFrame:
        """
        读取文件，自动检测格式
        
        Args:
            file_path: 文件路径
            sheet_name: Sheet名称（仅Excel有效）
            **kwargs: 传递给pandas的额外参数
            
        Returns:
            pd.DataFrame
        """
        file_path = Path(file_path)
        if not file_path.exists():
            raise FileNotFoundError(f"文件不存在: {file_path}")
        
        file_format = self.detect_format(str(file_path))
        logger.info(f"读取文件: {file_path.name} (格式: {file_format})")
        
        if file_format == 'csv':
            return self._read_csv(file_path, **kwargs)
        else:
            return self._read_excel(file_path, sheet_name, **kwargs)
    
    def _read_csv(self, file_path: Path, **kwargs) -> pd.DataFrame:
        """读取CSV文件，自动处理编码"""
        csv_config = self.config.get('csv', {})
        encodings = csv_config.get('encoding_priority', ['utf-8-sig', 'utf-8', 'gbk'])
        
        # 合并参数
        read_kwargs = {
            'low_memory': csv_config.get('low_memory', False),
            'dtype': str,  # 默认按字符串读取，避免类型推断问题
        }
        read_kwargs.update(kwargs)
        
        # 按编码优先级尝试
        last_error = None
        for encoding in encodings:
            try:
                df = pd.read_csv(file_path, encoding=encoding, **read_kwargs)
                logger.info(f"  使用编码: {encoding}, 行数: {len(df)}, 列数: {len(df.columns)}")
                return self._process_dataframe(df)
            except UnicodeDecodeError as e:
                last_error = e
                continue
            except Exception as e:
                logger.error(f"读取CSV失败 ({encoding}): {e}")
                raise
        
        raise last_error or Exception(f"无法读取CSV文件，尝试的编码: {encodings}")
    
    def _read_excel(self, file_path: Path, sheet_name: str = None, **kwargs) -> pd.DataFrame:
        """读取Excel文件"""
        excel_config = self.config.get('excel', {})
        
        read_kwargs = {
            'engine': 'openpyxl',
            'dtype': str,  # 默认按字符串读取
        }
        if sheet_name is not None:
            read_kwargs['sheet_name'] = sheet_name
        read_kwargs.update(kwargs)
        
        df = pd.read_excel(file_path, **read_kwargs)
        logger.info(f"  Sheet: {sheet_name or '默认'}, 行数: {len(df)}, 列数: {len(df.columns)}")
        
        return self._process_dataframe(df)
    
    def read_all_sheets(self, file_path: str) -> Dict[str, pd.DataFrame]:
        """
        读取Excel的所有Sheet
        
        Args:
            file_path: Excel文件路径
            
        Returns:
            {sheet_name: dataframe} 字典
        """
        file_path = Path(file_path)
        if not file_path.exists():
            raise FileNotFoundError(f"文件不存在: {file_path}")
        
        logger.info(f"读取所有Sheet: {file_path.name}")
        
        # 先获取所有Sheet名称
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        wb.close()
        
        result = {}
        for sheet_name in sheet_names:
            try:
                result[sheet_name] = self.read_file(str(file_path), sheet_name=sheet_name)
            except Exception as e:
                logger.warning(f"读取Sheet失败: {sheet_name}, 错误: {e}")
        
        logger.info(f"  成功读取 {len(result)} 个Sheet")
        return result
    
    def _process_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """处理DataFrame：类型转换、数据清洗等"""
        cleaning_config = self.config.get('cleaning', {})
        type_config = self.config.get('type_conversion', {})
        
        # 移除空行
        if cleaning_config.get('remove_empty_rows', True):
            original_len = len(df)
            df = df.dropna(how='all').reset_index(drop=True)
            if len(df) < original_len:
                logger.debug(f"  移除 {original_len - len(df)} 行空行")
        
        # 去除字符串首尾空格
        if cleaning_config.get('strip_strings', True):
            for col in df.select_dtypes(include=['object']).columns:
                df[col] = df[col].apply(
                    lambda x: x.strip() if isinstance(x, str) else x
                )
        
        # 类型转换
        if type_config.get('enabled', True):
            df = self._convert_data_types(df, type_config)
        
        return df
    
    def _convert_data_types(self, df: pd.DataFrame, type_config: Dict) -> pd.DataFrame:
        """自动转换数据类型"""
        date_keywords = type_config.get('date_columns', [])
        numeric_keywords = type_config.get('numeric_columns', [])
        force_string_keywords = type_config.get('force_string_columns', [])
        
        for col in df.columns:
            col_str = str(col)
            
            # 检查是否需要强制为字符串
            if any(keyword in col_str for keyword in force_string_keywords):
                df[col] = df[col].astype(str).replace('nan', np.nan)
                continue
            
            # 尝试日期转换
            if any(keyword in col_str for keyword in date_keywords):
                try:
                    df[col] = pd.to_datetime(df[col], errors='coerce')
                    continue
                except:
                    pass
            
            # 尝试数值转换
            if any(keyword in col_str for keyword in numeric_keywords):
                try:
                    df[col] = pd.to_numeric(df[col], errors='coerce')
                    continue
                except:
                    pass
        
        return df
    
    def write_excel(self, file_path: str, data: Dict[str, pd.DataFrame], 
                    template_path: str = None, **kwargs):
        """
        写入Excel文件
        
        Args:
            file_path: 输出文件路径
            data: {sheet_name: dataframe} 字典
            template_path: 可选模板路径（使用模板时会保留原有格式）
        """
        file_path = Path(file_path)
        file_path.parent.mkdir(parents=True, exist_ok=True)
        
        if template_path and Path(template_path).exists():
            # 使用模板写入
            self._write_with_template(file_path, data, template_path)
        else:
            # 直接写入
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                for sheet_name, df in data.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        logger.info(f"文件已保存: {file_path}")
    
    def _write_with_template(self, output_path: Path, data: Dict[str, pd.DataFrame], 
                             template_path: str):
        """使用模板写入数据（保留格式）"""
        wb = openpyxl.load_workbook(template_path)
        
        for sheet_name, df in data.items():
            if sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(sheet_name)
            else:
                ws = wb[sheet_name]
            
            # 写入数据（从第2行开始，假设第1行是表头）
            for row_idx, row_data in enumerate(df.itertuples(index=False), start=2):
                for col_idx, value in enumerate(row_data, start=1):
                    if pd.isna(value):
                        continue
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = value
        
        wb.save(str(output_path))
    
    @staticmethod
    def optimize_for_large_data(df: pd.DataFrame) -> pd.DataFrame:
        """
        优化大内存占用的DataFrame
        
        Args:
            df: 原始DataFrame
            
        Returns:
            优化后的DataFrame
        """
        optimized = df.copy()
        
        # 优化 object 类型为 category
        for col in optimized.select_dtypes(include=['object']).columns:
            num_unique = optimized[col].nunique()
            num_total = len(optimized)
            
            if num_unique / num_total < 0.5:  # 唯一值少于50%时优化
                optimized[col] = optimized[col].astype('category')
        
        # 下浮数值类型
        for col in optimized.select_dtypes(include=['float64']).columns:
            optimized[col] = pd.to_numeric(optimized[col], downcast='float')
        
        for col in optimized.select_dtypes(include=['int64']).columns:
            optimized[col] = pd.to_numeric(optimized[col], downcast='integer')
        
        original_mem = df.memory_usage(deep=True).sum() / 1024 / 1024
        optimized_mem = optimized.memory_usage(deep=True).sum() / 1024 / 1024
        
        logger.info(f"内存优化: {original_mem:.2f} MB -> {optimized_mem:.2f} MB "
                   f"(节省 {100 - (optimized_mem/original_mem*100):.1f}%)")
        
        return optimized
    
    def get_workbook_info(self, file_path: str) -> Dict[str, Any]:
        """
        获取工作簿基本信息
        
        Args:
            file_path: Excel文件路径
            
        Returns:
            包含工作簿信息的字典
        """
        file_path = Path(file_path)
        wb = openpyxl.load_workbook(file_path, read_only=True)
        
        info = {
            'sheet_count': len(wb.sheetnames),
            'sheet_names': wb.sheetnames,
            'sheets': {}
        }
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            info['sheets'][sheet_name] = {
                'max_row': ws.max_row,
                'max_column': ws.max_column,
                'dimensions': ws.dimensions
            }
        
        wb.close()
        return info
