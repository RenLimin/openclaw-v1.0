#!/usr/bin/env python3
"""
样式批量应用模块
================
从模板复制样式到目标文件，支持按行/列批量设置样式
"""

import logging
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
import openpyxl
from openpyxl.styles import PatternFill, Border, Font, Alignment, Side
from openpyxl.utils import get_column_letter
import copy

logger = logging.getLogger(__name__)


class StyleApplicator:
    """
    样式批量应用类
    
    功能:
    - 从模板复制样式
    - 按行批量应用样式
    - 按列批量应用样式
    - 按区域批量应用样式
    """
    
    def __init__(self, template_path: str = None):
        """
        初始化样式应用器
        
        Args:
            template_path: 模板文件路径（可选）
        """
        self.template_path = Path(template_path) if template_path else None
        self.template_wb: Optional[openpyxl.Workbook] = None
        self._style_cache: Dict[str, Any] = {}
        
        if self.template_path and self.template_path.exists():
            self.template_wb = openpyxl.load_workbook(
                self.template_path,
                data_only=False
            )
            logger.info(f"已加载模板: {self.template_path.name}")
    
    def copy_style_from_template(self, template_sheet: str, target_wb, target_sheet: str,
                                 row_range: Tuple[int, int] = None, 
                                 col_range: Tuple[int, int] = None,
                                 start_row: int = 1,
                                 start_col: int = 1):
        """
        从模板复制样式到目标文件
        
        Args:
            template_sheet: 模板Sheet名
            target_wb: 目标工作簿
            target_sheet: 目标Sheet名
            row_range: 模板中行范围 (start, end)，None表示全部
            col_range: 模板中列范围 (start, end)，None表示全部
            start_row: 目标起始行
            start_col: 目标起始列
        """
        if not self.template_wb:
            logger.warning("未加载模板，无法复制样式")
            return
        
        if template_sheet not in self.template_wb.sheetnames:
            logger.warning(f"模板Sheet不存在: {template_sheet}")
            return
        
        template_ws = self.template_wb[template_sheet]
        
        if target_sheet not in target_wb.sheetnames:
            target_ws = target_wb.create_sheet(target_sheet)
        else:
            target_ws = target_wb[target_sheet]
        
        # 确定复制范围
        t_start_row, t_end_row = row_range or (1, template_ws.max_row)
        t_start_col, t_end_col = col_range or (1, template_ws.max_column)
        
        logger.info(f"复制样式: 模板 {template_sheet} ({t_start_row}-{t_end_row}行, {t_start_col}-{t_end_col}列) "
                   f"-> 目标 {target_sheet} (从 {start_row},{start_col} 开始)")
        
        # 复制样式
        copied_count = 0
        for row_offset in range(t_end_row - t_start_row + 1):
            for col_offset in range(t_end_col - t_start_col + 1):
                t_row = t_start_row + row_offset
                t_col = t_start_col + col_offset
                
                target_row = start_row + row_offset
                target_col = start_col + col_offset
                
                t_cell = template_ws.cell(row=t_row, column=t_col)
                target_cell = target_ws.cell(row=target_row, column=target_col)
                
                # 复制样式
                self._copy_cell_style(t_cell, target_cell)
                copied_count += 1
        
        logger.info(f"  已复制 {copied_count} 个单元格样式")
    
    def _copy_cell_style(self, source_cell, target_cell):
        """复制单个单元格的样式"""
        # 填充色
        if source_cell.fill:
            target_cell.fill = copy.copy(source_cell.fill)
        
        # 字体
        if source_cell.font:
            target_cell.font = copy.copy(source_cell.font)
        
        # 边框
        if source_cell.border:
            target_cell.border = copy.copy(source_cell.border)
        
        # 对齐
        if source_cell.alignment:
            target_cell.alignment = copy.copy(source_cell.alignment)
        
        # 数字格式
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format
    
    def apply_row_style(self, ws, row_num: int, style_template=None, style_dict: Dict = None):
        """
        应用行样式
        
        Args:
            ws: Worksheet对象
            row_num: 行号
            style_template: 模板单元格样式（优先）
            style_dict: 样式字典 {fill, font, border, alignment, number_format}
        """
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_num, column=col)
            
            if style_template:
                self._copy_cell_style(style_template, cell)
            elif style_dict:
                self._apply_style_dict(cell, style_dict)
    
    def apply_column_style(self, ws, col_num: int, style_template=None, style_dict: Dict = None):
        """
        应用列样式
        
        Args:
            ws: Worksheet对象
            col_num: 列号
            style_template: 模板单元格样式（优先）
            style_dict: 样式字典
        """
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_num)
            
            if style_template:
                self._copy_cell_style(style_template, cell)
            elif style_dict:
                self._apply_style_dict(cell, style_dict)
    
    def apply_range_style(self, ws, start_row: int, end_row: int, 
                          start_col: int, end_col: int,
                          style_template=None, style_dict: Dict = None):
        """
        应用区域样式
        
        Args:
            ws: Worksheet对象
            start_row: 起始行
            end_row: 结束行
            start_col: 起始列
            end_col: 结束列
            style_template: 模板单元格样式
            style_dict: 样式字典
        """
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)
                
                if style_template:
                    self._copy_cell_style(style_template, cell)
                elif style_dict:
                    self._apply_style_dict(cell, style_dict)
        
        logger.info(f"已应用样式到区域: {get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}")
    
    def _apply_style_dict(self, cell, style_dict: Dict):
        """应用样式字典到单元格"""
        if 'fill' in style_dict:
            cell.fill = style_dict['fill']
        
        if 'font' in style_dict:
            cell.font = style_dict['font']
        
        if 'border' in style_dict:
            cell.border = style_dict['border']
        
        if 'alignment' in style_dict:
            cell.alignment = style_dict['alignment']
        
        if 'number_format' in style_dict:
            cell.number_format = style_dict['number_format']
    
    @staticmethod
    def create_fill(color: str, pattern: str = 'solid') -> PatternFill:
        """
        创建填充样式
        
        Args:
            color: RGB颜色值，如 "FFCCCCCC"
            pattern: 填充模式，默认 'solid'
            
        Returns:
            PatternFill对象
        """
        return PatternFill(start_color=color, end_color=color, fill_type=pattern)
    
    @staticmethod
    def create_font(name: str = '宋体', size: int = 11, bold: bool = False,
                    italic: bool = False, color: str = None) -> Font:
        """
        创建字体样式
        
        Args:
            name: 字体名称
            size: 字号
            bold: 是否加粗
            italic: 是否斜体
            color: 字体颜色RGB
            
        Returns:
            Font对象
        """
        kwargs = {
            'name': name,
            'size': size,
            'bold': bold,
            'italic': italic
        }
        if color:
            kwargs['color'] = color
        
        return Font(**kwargs)
    
    @staticmethod
    def create_border(style: str = 'thin', color: str = 'FF000000') -> Border:
        """
        创建边框样式
        
        Args:
            style: 边框样式: 'thin', 'thick', 'medium', 'dashed', 等
            color: 边框颜色RGB
            
        Returns:
            Border对象
        """
        side = Side(style=style, color=color)
        return Border(left=side, right=side, top=side, bottom=side)
    
    @staticmethod
    def create_alignment(horizontal: str = 'center', vertical: str = 'center',
                         wrap_text: bool = False) -> Alignment:
        """
        创建对齐样式
        
        Args:
            horizontal: 水平对齐: 'left', 'center', 'right'
            vertical: 垂直对齐: 'top', 'center', 'bottom'
            wrap_text: 是否自动换行
            
        Returns:
            Alignment对象
        """
        return Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap_text)
    
    def apply_alternating_row_colors(self, ws, start_row: int = 2, end_row: int = None,
                                     color1: str = 'FFFFFFFF', color2: str = 'FFF5F5F5'):
        """
        应用隔行变色
        
        Args:
            ws: Worksheet对象
            start_row: 起始行
            end_row: 结束行，None表示到最后一行
            color1: 奇数行颜色
            color2: 偶数行颜色
        """
        if end_row is None:
            end_row = ws.max_row
        
        fill1 = self.create_fill(color1)
        fill2 = self.create_fill(color2)
        
        for row in range(start_row, end_row + 1):
            fill = fill1 if (row - start_row) % 2 == 0 else fill2
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.fill = fill
        
        logger.info(f"已应用隔行变色: 行 {start_row}-{end_row}")
    
    def apply_table_borders(self, ws, start_row: int = 1, end_row: int = None,
                            start_col: int = None, end_col: int = None,
                            border_style: str = 'thin'):
        """
        应用表格边框
        
        Args:
            ws: Worksheet对象
            start_row: 起始行
            end_row: 结束行
            start_col: 起始列
            end_col: 结束列
            border_style: 边框样式
        """
        if end_row is None:
            end_row = ws.max_row
        if start_col is None:
            start_col = 1
        if end_col is None:
            end_col = ws.max_column
        
        border = self.create_border(style=border_style)
        
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = border
        
        logger.info(f"已应用表格边框: {get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}")
    
    def apply_header_style(self, ws, header_row: int = 1, bold: bool = True,
                           bg_color: str = 'FFD9D9D9', font_size: int = 11):
        """
        应用表头样式
        
        Args:
            ws: Worksheet对象
            header_row: 表头行号
            bold: 是否加粗
            bg_color: 背景色
            font_size: 字体大小
        """
        style = {
            'fill': self.create_fill(bg_color),
            'font': self.create_font(bold=bold, size=font_size),
            'alignment': self.create_alignment()
        }
        
        self.apply_row_style(ws, header_row, style_dict=style)
        
        logger.info(f"已应用表头样式: 第 {header_row} 行")
    
    def auto_fit_columns(self, ws, min_width: int = 10, max_width: int = 50):
        """
        自动调整列宽
        
        Args:
            ws: Worksheet对象
            min_width: 最小宽度
            max_width: 最大宽度
        """
        for col in range(1, ws.max_column + 1):
            max_length = 0
            column_letter = get_column_letter(col)
            
            for row in range(1, ws.max_row + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value:
                    # 粗略估计显示宽度
                    length = len(str(cell.value)) * 1.2
                    if length > max_length:
                        max_length = length
            
            # 设置列宽
            width = max(min_width, min(max_length, max_width))
            ws.column_dimensions[column_letter].width = width
            
        logger.info(f"已自动调整 {ws.max_column} 列的宽度")
    
    def copy_sheet_style(self, source_ws, target_ws, copy_row_heights: bool = True,
                         copy_col_widths: bool = True):
        """
        复制整个Sheet的样式
        
        Args:
            source_ws: 源Worksheet
            target_ws: 目标Worksheet
            copy_row_heights: 是否复制行高
            copy_col_widths: 是否复制列宽
        """
        # 复制单元格样式
        for row in range(1, source_ws.max_row + 1):
            for col in range(1, source_ws.max_column + 1):
                source_cell = source_ws.cell(row=row, column=col)
                target_cell = target_ws.cell(row=row, column=col)
                self._copy_cell_style(source_cell, target_cell)
        
        # 复制行高
        if copy_row_heights:
            for row in range(1, source_ws.max_row + 1):
                if source_ws.row_dimensions[row].height:
                    target_ws.row_dimensions[row].height = source_ws.row_dimensions[row].height
        
        # 复制列宽
        if copy_col_widths:
            for col in range(1, source_ws.max_column + 1):
                col_letter = get_column_letter(col)
                if col_letter in source_ws.column_dimensions:
                    target_ws.column_dimensions[col_letter].width = \
                        source_ws.column_dimensions[col_letter].width
        
        logger.info(f"已复制Sheet样式: {source_ws.title} -> {target_ws.title}")
    
    def close(self):
        """关闭模板工作簿"""
        if self.template_wb:
            self.template_wb.close()
    
    def __enter__(self):
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()
