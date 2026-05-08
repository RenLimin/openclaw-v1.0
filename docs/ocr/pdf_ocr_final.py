#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
PDF OCR 合同条款拆分 - 最终交付版
"""

import re
import yaml
from pathlib import Path
from typing import List, Dict


# ========== 颜色输出 ==========
class Colors:
    RED = '\033[91m'
    GREEN = '\033[92m'
    YELLOW = '\033[93m'
    CYAN = '\033[96m'
    ENDC = '\033[0m'


# ========== OCR 引擎 v3.0 ==========
OCR_CORRECTION_DICT = {
    '安时': '安卓', '内起三': '1、', '内起': '1、', '起三': '1、', '文付': '支付',
    '付亚': '付款', 'BiaA': '', 'fcoFBSH': '', 'ARMAS': '', 'Siannltc': '', 'FBSH': '',
    'fco': '', '®': '', '@': '', '违造': '违约', '造约': '违约', '违备': '违约',
    '尝还': '偿还', '尝付': '偿付', '保正': '保证', '负任': '责任', '责仁': '责任',
    '叉方': '双方', '甲力': '甲方', '乙力': '乙方', '丙力': '丙方', '合问': '合同',
    '合司': '合同', '条敦': '条款', '条欺': '条款', '条软': '条款', '权力': '权利',
    '叉利': '权利', '艾务': '义务', '又务': '义务', '陪偿': '赔偿', '培偿': '赔偿',
    '滞钠金': '滞纳金', '滞拿金': '滞纳金', '人民市': '人民币', '人民而': '人民币',
    '帮3': '3、', '帮': '',
}


def is_chinese_char(c: str) -> bool:
    return '\u4e00' <= c <= '\u9fff'


def is_valid_char(c: str) -> bool:
    if c.isspace(): return True
    if c.isalnum(): return True
    if c in '，。、；：“”‘’（）【】《》！？,.?!()[]""\'<>%¥$&+-=\\': return True
    if is_chinese_char(c): return True
    return False


def filter_garbage(text: str) -> str:
    return ''.join(c for c in text if is_valid_char(c))


def remove_table_noise(text: str) -> str:
    lines = text.split('\n')
    cleaned = []
    for line in lines:
        line = line.strip()
        if not line: continue
        special = sum(1 for c in line if not c.isalnum() and not is_chinese_char(c))
        if len(line) > 0 and special / len(line) > 0.6: continue
        if re.match(r'^[A-Z]{4,15}$', line): continue
        cleaned.append(line)
    return '\n'.join(cleaned)


def correct_ocr(text: str) -> str:
    for wrong, right in OCR_CORRECTION_DICT.items():
        text = text.replace(wrong, right)
    return text


def merge_paragraphs(lines: List[str]) -> List[str]:
    if not lines: return []
    paragraphs = []
    current = lines[0]
    end_punc = {'。', '！', '？', '；', '：', '”', '）', '】', '》'}
    start_kw = {'第', '一', '二', '三', '四', '五', '六', '七', '八', '九', '十',
                '甲', '乙', '丙', '丁', '（', '(', '【', '《', '1', '2', '3', '4', '5', '6', '7', '8', '9', '0'}
    
    for line in lines[1:]:
        line = line.strip()
        if not line: continue
        prev_ended = current and current[-1] in end_punc
        is_new = len(line) >= 1 and line[0] in start_kw
        if prev_ended and is_new:
            paragraphs.append(current)
            current = line
        else:
            current += line
    if current: paragraphs.append(current)
    return paragraphs


def ocr_pdf(pdf_path: Path) -> str:
    print(f"\n{Colors.CYAN}{'='*60}{Colors.ENDC}")
    print(f"  📜 PDF OCR 合同条款拆分 v3.0")
    print(f"{Colors.CYAN}{'='*60}{Colors.ENDC}")
    
    try:
        import fitz
        doc = fitz.open(str(pdf_path))
        total, paras = 0, []
        for page in doc:
            text = page.get_text()
            total += len(text.strip())
            for p in text.split('\n'):
                if p.strip() and len(p.strip()) > 2: paras.append(p.strip())
        if total / len(doc) > 100:
            print(f"  ✅ 检测为【文本版 PDF】")
            return '\n'.join(paras)
    except: pass
    
    print(f"  📷 检测到扫描版 PDF，启动 OCR ...")
    from pdf2image import convert_from_path
    import pytesseract
    
    dpi = 450
    print(f"  🖨️  PDF 转图片 (DPI={dpi}) ...")
    pages = convert_from_path(str(pdf_path), dpi=dpi)
    print(f"  ✅ 共 {len(pages)} 页")
    
    print(f"  🤖  OCR 识别中（三模式融合）...")
    configs = [
        f'--oem 3 --psm 6 -l chi_sim --dpi {dpi}',
        f'--oem 3 --psm 11 -l chi_sim --dpi {dpi}',
        f'--oem 3 --psm 3 -l chi_sim --dpi {dpi}',
    ]
    
    all_lines = []
    for i, page in enumerate(pages, 1):
        print(f"  🔍  识别第 {i}/{len(pages)} 页 ...", end='\r')
        best, max_valid = [], 0
        for cfg in configs:
            text = pytesseract.image_to_string(page, config=cfg)
            lines = [l.strip() for l in text.split('\n') if l.strip()]
            valid = sum(len(l) for l in lines if any(is_chinese_char(c) for c in l))
            if valid > max_valid: max_valid = valid; best = lines
        all_lines.extend(best)
    
    print(f"\n  🧹  深度后处理 ...")
    raw = '\n'.join(all_lines)
    text = filter_garbage(raw)
    text = remove_table_noise(text)
    text = correct_ocr(text)
    merged = merge_paragraphs(text.split('\n'))
    
    final = '\n'.join(merged)
    print(f"  ✅ 完成: {len(merged)} 段落, {len(final)} 字符")
    return final


# ========== 条款拆分/分类/风险识别 ==========
def split_clauses(full_text: str) -> List[Dict]:
    print(f"\n{Colors.CYAN}✂️  条款智能拆分中...{Colors.ENDC}")
    patterns = [r'^\s*第[一二三四五六七八九十\d]+[条章节款]', r'^\s*\d+\.\s',
                r'^\s*\d+\.\d+\s', r'^\s*[（(][一二三四五六七八九十\d]+[）)]\s*',
                r'^\s*[一二三四五六七八九十]{1,2}[、．.]\s*']
    combined = re.compile('|'.join(patterns))
    
    lines = full_text.split('\n')
    clauses, current = [], []
    for line in lines:
        line = line.strip()
        if not line or len(line) < 3: continue
        if combined.match(line) and current:
            clauses.append({'id': len(clauses)+1, 'content': '\n'.join(current), 'length': sum(len(c) for c in current)})
            current = [line]
        else: current.append(line)
    if current: clauses.append({'id': len(clauses)+1, 'content': '\n'.join(current), 'length': sum(len(c) for c in current)})
    print(f"  ✅ 拆分完成: {len(clauses)} 个条款")
    return clauses


def classify_clauses(clauses: List[Dict], rules: List[Dict]) -> List[Dict]:
    print(f"\n{Colors.CYAN}🏷️  条款分类中...{Colors.ENDC}")
    cats = {}
    for idx, clause in enumerate(clauses):
        content, best_score, best_prio, matched = clause['content'], -1, 0, None
        is_early = idx < 4
        
        force = None
        if '[表格' in content or '设备清单' in content or '配置清单' in content:
            for r in rules:
                if r.get('name') == '标的条款': force = r; break
        elif '人民币' in content and '元整' in content and '违约金' not in content:
            for r in rules:
                if r.get('name') == '价格条款': force = r; break
        
        if force:
            clause['category'] = force.get('name', '')
            cats[clause['category']] = cats.get(clause['category'], 0) + 1
            continue
        
        for rule in rules:
            name = rule.get('name', '')
            exclude = rule.get('exclude_keywords', [])
            if any(kw in content for kw in exclude): continue
            score = sum(3 for feat in rule.get('core_features', []) if feat in content)
            score += sum(1 for kw in rule.get('keywords', []) if kw in content)
            if is_early and name == '基本信息': score += 3
            if is_early and name in ['价格条款', '违约责任']: score -= 2
            priority = rule.get('priority', 0)
            if (priority > best_prio) or (priority == best_prio and score > best_score):
                best_score, best_prio, matched = score, priority, rule
        
        clause['category'] = matched.get('name', '待分类') if matched else '待分类'
        cats[clause['category']] = cats.get(clause['category'], 0) + 1
    
    print(f"  {Colors.CYAN}📊 分类统计:{Colors.ENDC}")
    for cat, cnt in sorted(cats.items()):
        print(f"    {cat}: {cnt} 条")
    return clauses


def identify_risks(clauses: List[Dict], risk_cfg: Dict) -> List[Dict]:
    print(f"\n{Colors.CYAN}⚠️  风险识别中...{Colors.ENDC}")
    high_kws = risk_cfg.get('high_risk', {}).get('keywords', [])
    medium_kws = risk_cfg.get('medium_risk', {}).get('keywords', [])
    low_kws = risk_cfg.get('low_risk', {}).get('keywords', [])
    risks = {'high': 0, 'medium': 0, 'low': 0}
    
    for clause in clauses:
        content, risk_lvl, risk_kws = clause['content'], "✅ 无风险", []
        for kw in high_kws:
            if kw in content: risk_lvl = "🔴 高风险"; risk_kws.append(kw); break
        if risk_lvl == "✅ 无风险":
            for kw in medium_kws:
                if kw in content: risk_lvl = "🟡 中风险"; risk_kws.append(kw); break
        if risk_lvl == "✅ 无风险":
            for kw in low_kws:
                if kw in content: risk_lvl = "🟢 低风险"; risk_kws.append(kw); break
        clause['risk_level'], clause['risk_keywords'] = risk_lvl, ', '.join(risk_kws)
        if '高风险' in risk_lvl: risks['high'] += 1
        elif '中风险' in risk_lvl: risks['medium'] += 1
        elif '低风险' in risk_lvl: risks['low'] += 1
    
    print(f"  🔴 高风险: {risks['high']} 条")
    print(f"  🟡 中风险: {risks['medium']} 条")
    print(f"  🟢 低风险: {risks['low']} 条")
    return clauses


def export_excel(clauses: List[Dict], output: Path):
    print(f"\n{Colors.CYAN}📊 导出 Excel...{Colors.ENDC}")
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    wb = openpyxl.Workbook()
    h_font = Font(bold=True, size=12, color="FFFFFF")
    h_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    r_high_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    r_medium_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    
    ws1 = wb.active
    ws1.title = "全部条款"
    for col, h in enumerate(['条款编号', '条款分类', '风险等级', '风险关键词', '条款内容', '字数'], 1):
        c = ws1.cell(row=1, column=col, value=h)
        c.font = h_font; c.fill = h_fill
    for row, clause in enumerate(clauses, 2):
        ws1.cell(row=row, column=1, value=clause['id'])
        ws1.cell(row=row, column=2, value=clause['category'])
        ws1.cell(row=row, column=3, value=clause['risk_level'])
        ws1.cell(row=row, column=4, value=clause['risk_keywords'])
        ws1.cell(row=row, column=5, value=clause['content'])
        ws1.cell(row=row, column=6, value=clause['length'])
        if '高风险' in clause['risk_level']: ws1.cell(row=row, column=3).fill = r_high_fill
        elif '中风险' in clause['risk_level']: ws1.cell(row=row, column=3).fill = r_medium_fill
    ws1.column_dimensions['A'].width = 10
    ws1.column_dimensions['B'].width = 14
    ws1.column_dimensions['C'].width = 12
    ws1.column_dimensions['D'].width = 20
    ws1.column_dimensions['E'].width = 100
    
    ws2 = wb.create_sheet("分类视图")
    cat_list = sorted(list(set(c['category'] for c in clauses)))
    row = 1
    for cat in cat_list:
        ws2.cell(row=row, column=1, value=cat).font = Font(bold=True, size=14)
        row += 1
        for clause in [c for c in clauses if c['category'] == cat]:
            ws2.cell(row=row, column=1, value=f"条款 {clause['id']}")
            ws2.cell(row=row, column=2, value=clause['risk_level'])
            ws2.cell(row=row, column=3, value=clause['content'][:200])
            row += 1
        row += 1
    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 120
    
    ws3 = wb.create_sheet("风险条款汇总")
    for col, h in enumerate(['条款编号', '风险等级', '风险关键词', '条款内容'], 1):
        c = ws3.cell(row=1, column=col, value=h)
        c.font = h_font; c.fill = h_fill
    risk_clauses = [c for c in clauses if '高风险' in c['risk_level'] or '中风险' in c['risk_level']]
    for row, clause in enumerate(risk_clauses, 2):
        ws3.cell(row=row, column=1, value=clause['id'])
        ws3.cell(row=row, column=2, value=clause['risk_level'])
        ws3.cell(row=row, column=3, value=clause['risk_keywords'])
        ws3.cell(row=row, column=4, value=clause['content'])
        if '高风险' in clause['risk_level']: ws3.cell(row=row, column=2).fill = r_high_fill
        elif '中风险' in clause['risk_level']: ws3.cell(row=row, column=2).fill = r_medium_fill
    ws3.column_dimensions['A'].width = 10
    ws3.column_dimensions['B'].width = 12
    ws3.column_dimensions['C'].width = 30
    ws3.column_dimensions['D'].width = 120
    
    ws4 = wb.create_sheet("OCR原始文本")
    ws4.cell(row=1, column=1, value="OCR识别原始文本（用于核对）").font = h_font
    ws4.cell(row=1, column=1).fill = h_fill
    ws4.cell(row=2, column=1, value='\n'.join(c['content'] for c in clauses)[:32767])
    ws4.column_dimensions['A'].width = 120
    
    wb.save(str(output))
    print(f"  ✅ 已导出: {output}")


def main():
    pdf_path = Path("/Users/bangcle/Downloads/contract/XSZS2603090130北京聚信得仁.pdf")
    output = Path("/Users/bangcle/Downloads/contract/XSZS2603090130_OCR_最终版.xlsx")
    cfg_dir = Path("/Users/bangcle/.openclaw/workspace/skills/contract-clause-split/config")
    
    with open(cfg_dir / "classification-rules.yaml", 'r', encoding='utf-8') as f:
        class_cfg = yaml.safe_load(f)
    with open(cfg_dir / "risk-keywords.yaml", 'r', encoding='utf-8') as f:
        risk_cfg = yaml.safe_load(f)
    rules = class_cfg.get('classification_rules', [])
    
    full_text = ocr_pdf(pdf_path)
    clauses = split_clauses(full_text)
    clauses = classify_clauses(clauses, rules)
    clauses = identify_risks(clauses, risk_cfg)
    export_excel(clauses, output)
    
    print(f"\n{Colors.GREEN}🎉 PDF OCR 最终版已完成！{Colors.ENDC}")
    print(f"   📂 输出文件: {output}")
    import subprocess
    subprocess.run(['open', str(output)])


if __name__ == "__main__":
    main()
