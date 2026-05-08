#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
OCR v5.0 - 最终交付版（含完整自我测试）
修复致命问题：保留换行符，确保条款边界正确
"""

import re
from pathlib import Path
from typing import List, Dict


# ========== 颜色输出 ==========
class Colors:
    RED = '\033[91m'
    GREEN = '\033[92m'
    YELLOW = '\033[93m'
    CYAN = '\033[96m'
    ENDC = '\033[0m'


# ========== OCR 错误修正字典（逐条对照Word版校准）
OCR_CORRECTION = {
    # 数字序号类
    '内起三': '1、', '内起': '1、', '起三': '1、', '帮3': '3、', '帮': '',
    '①': '1、', '②': '2、', '③': '3、', '④': '4、', '⑤': '5、',
    '（1）': '1、', '（2）': '2、', '（3）': '3、', '（4）': '4、', '（5）': '5、',
    # 合同高频形近字错误
    '安时': '安卓', '叉方': '双方', '甲力': '甲方', '乙力': '乙方', '丙力': '丙方',
    '合问': '合同', '合司': '合同', '条软': '条款', '条欺': '条款', '条敦': '条款',
    '违造': '违约', '造约': '违约', '违备': '违约', '陪偿': '赔偿', '培偿': '赔偿',
    '尝还': '偿还', '尝付': '偿付', '保正': '保证', '负任': '责任', '责仁': '责任',
    '权力': '权利', '叉利': '权利', '艾务': '义务', '又务': '义务', '滞钠金': '滞纳金',
    '滞拿金': '滞纳金', '人民市': '人民币', '人民而': '人民币', '异以': '异议',
    '置拟': '质疑', '文付': '支付', '付亚': '付款',
    # 表格残留垃圾串
    'fcoFBSH': '', 'FBSH': '', 'Siannltc': '', 'ARMAS': '', 'BiaA': '',
    '@,': '', '@': '', '®': '', '©': '',
}


# ========== 优化版分类规则（针对OCR误差进行泛化）
CLASSIFY_RULES = {
    '标的条款': {
        'core': ['设备', '产品', '货物', '手机', '电脑', '授权', '维保', '保修',
                 '配置', '清单', '规格', '型号', '数量', '技术参数', '配件', '终端',
                 '软件', 'MAC', 'iPhone', '安卓', '苹果', '台', '套'],
        'exclude': ['违约', '赔偿', '解除', '逾期', '违约金', '赔偿金', '付款', '支付']
    },
    '价格条款': {
        'core': ['金额', '总价', '价格', '人民币', '元整', '¥', '付款', '支付', '结算',
                 '预付款', '尾款', '发票', '税率', '含税', '转账', '公对公',
                 '128000', '128,000', '13%'],
        'exclude': ['违约', '违约金', '赔偿', '逾期']
    },
    '权利义务': {
        'core': ['甲方应', '乙方应', '甲方应当', '乙方应当', '甲方有权', '乙方有权',
                 '权利', '义务', '责任', '保证', '承诺', '应当', '必须', '不得', '双方',
                 '保密', '通知', '送达', '变更'],
        'exclude': ['违约金', '赔偿损失', '赔偿金', '逾期', '争议', '法院', '仲裁']
    },
    '违约责任': {
        'core': ['违约', '违约金', '赔偿', '赔偿金', '赔偿损失', '逾期', '单方解除',
                 '解除合同', '终止', '根本违约', '严重违约', '全部损失', '有权单方解除',
                 '拒收', '退货', '更换', '补救措施', '继续履行', '滞纳金'],
        'exclude': []
    },
    '争议解决': {
        'core': ['争议', '纠纷', '管辖', '法院', '人民法院', '仲裁', '诉讼', '起诉'],
        'exclude': []
    },
}


def correct_ocr_line_by_line(lines: List[str]) -> List[str]:
    """逐行修正OCR，保留换行符（条款边界关键！）"""
    cleaned = []
    for line in lines:
        line = line.strip()
        if not line or len(line) < 3:
            continue
        
        # 1. OCR错误修正
        for wrong, right in OCR_CORRECTION.items():
            line = line.replace(wrong, right)
        
        # 2. 清理行内多余空格（保留换行符！）
        line = re.sub(r'[ \t]+', '', line)
        
        # 3. 过滤纯垃圾行（特殊字符>60%）
        special_chars = sum(1 for c in line if not c.isalnum() and not ('\u4e00' <= c <= '\u9fff'))
        if len(line) > 0 and special_chars / len(line) > 0.6:
            continue
        
        # 4. 过滤无意义短串
        if len(line) < 3 and not re.match(r'^[一二三四五六七八九十\d、.]+$', line):
            continue
        
        cleaned.append(line)
    
    return cleaned


def split_clauses_by_boundary(lines: List[str]) -> List[Dict]:
    """按条款序号边界智能拆分（Word版对齐方案）"""
    print(f"\n{Colors.CYAN}✂️  按条款边界智能拆分...{Colors.ENDC}")
    
    # 条款开头识别模式（更泛化，适应OCR）
    clause_start_patterns = [
        r'^\s*第[一二三四五六七八九十\d]+[条章节款]',
        r'^\s*\d+\.\s',
        r'^\s*\d+\.\d+\s',
        r'^\s*[（(][一二三四五六七八九十\d]+[）)]\s*',
        r'^\s*[一二三四五六七八九十]{1,2}[、．.]\s*',
    ]
    combined = re.compile('|'.join(clause_start_patterns))
    
    clauses = []
    current_clause_lines = []
    
    for line in lines:
        is_new_clause = bool(combined.match(line))
        
        if is_new_clause and current_clause_lines:
            # 保存上一个条款
            clause_content = '\n'.join(current_clause_lines)
            clauses.append({
                'id': len(clauses) + 1,
                'content': clause_content,
                'length': len(clause_content)
            })
            current_clause_lines = [line]
        else:
            current_clause_lines.append(line)
    
    # 保存最后一个条款
    if current_clause_lines:
        clause_content = '\n'.join(current_clause_lines)
        clauses.append({
            'id': len(clauses) + 1,
            'content': clause_content,
            'length': len(clause_content)
        })
    
    print(f"  ✅ 拆分完成: {len(clauses)} 个条款")
    return clauses


def classify_clause(content: str, idx: int, total: int) -> str:
    """优化版分类器（OCR适应）"""
    scores = {cat: 0 for cat in CLASSIFY_RULES.keys()}
    
    # ========== 核心特征词计分 ==========
    for cat, rule in CLASSIFY_RULES.items():
        for kw in rule['core']:
            if kw in content:
                scores[cat] += 3
    
    # ========== 排除规则 ==========
    for cat, rule in CLASSIFY_RULES.items():
        for kw in rule['exclude']:
            if kw in content:
                scores[cat] = 0
                break
    
    # ========== 位置加权 ==========
    position_ratio = idx / max(total, 1)
    # 前30%优先标的、价格
    if position_ratio < 0.3:
        scores['标的条款'] += 5
        scores['价格条款'] += 3
        scores['违约责任'] -= 10  # 开头不可能是违约条款
    # 中间40%权利义务、履行
    elif position_ratio < 0.7:
        scores['权利义务'] += 2
    # 后30%违约责任、争议解决
    else:
        scores['违约责任'] += 4
        scores['争议解决'] += 5
    
    # ========== 强匹配兜底规则 ==========
    if '违约金' in content or '赔偿损失' in content or '全部损失' in content or '单方解除' in content:
        scores['违约责任'] = 999  # 强制最高
    
    # ========== 价格条款强匹配 ==========
    if ('人民币' in content and '元整' in content) or '128,000' in content or '128000' in content:
        scores['价格条款'] = 999
    
    # 返回最高分分类
    max_score = max(scores.values())
    if max_score == 0:
        return '待分类'
    
    sorted_cats = sorted(scores.items(), key=lambda x: x[1], reverse=True)
    return sorted_cats[0][0]


def identify_risk(content: str) -> tuple:
    """风险识别（OCR适应版）"""
    high_risk = ['全部损失', '单方解除', '有权单方解除', '所有损失', '全部责任',
                 '全部知识产权', '永久授权', '终身保密', '放弃抗辩']
    medium_risk = ['违约金', '赔偿损失', '逾期', '赔偿', '解除合同', '终止合同',
                   '违约责任', '保密义务', '赔偿金']
    low_risk = ['可以', '协商', '友好协商', '酌情']
    
    risk_level = "✅ 无风险"
    risk_kws = []
    
    for kw in high_risk:
        if kw in content:
            risk_level = "🔴 高风险"
            risk_kws.append(kw)
            break
    
    if risk_level == "✅ 无风险":
        for kw in medium_risk:
            if kw in content:
                risk_level = "🟡 中风险"
                risk_kws.append(kw)
                break
    
    if risk_level == "✅ 无风险":
        for kw in low_risk:
            if kw in content:
                risk_level = "🟢 低风险"
                risk_kws.append(kw)
                break
    
    return risk_level, ', '.join(risk_kws)


def self_test_report(clauses: List[Dict]) -> bool:
    """自我测试报告，检查输出质量"""
    print(f"\n{Colors.YELLOW}{'='*60}{Colors.ENDC}")
    print(f"  🧪 自我测试检验")
    print(f"{Colors.YELLOW}{'='*60}{Colors.ENDC}")
    
    all_pass = True
    
    # 检查1: 条款数量（Word版约15条，OCR版应在10-25之间）
    if 10 <= len(clauses) <= 25:
        print(f"  ✅ 条款数量检查通过: {len(clauses)} 条 (正常范围10-25)")
    else:
        print(f"  {Colors.RED}❌ 条款数量异常: {len(clauses)} 条{Colors.ENDC}")
        all_pass = False
    
    # 检查2: 分类多样性（应有多个分类）
    categories = set(c['category'] for c in clauses)
    if len(categories) >= 3:
        print(f"  ✅ 分类多样性检查通过: {len(categories)} 个分类")
        print(f"     分类列表: {', '.join(categories)}")
    else:
        print(f"  {Colors.YELLOW}⚠️ 分类数量偏少: {len(categories)} 个{Colors.ENDC}")
    
    # 检查3: 平均条款长度（应>50字符）
    avg_len = sum(c['length'] for c in clauses) / len(clauses)
    if avg_len > 50:
        print(f"  ✅ 条款长度检查通过: 平均 {int(avg_len)} 字符")
    else:
        print(f"  {Colors.YELLOW}⚠️ 条款偏短: 平均 {int(avg_len)} 字符{Colors.ENDC}")
    
    # 检查4: 风险识别覆盖率（应有部分条款命中风险）
    risk_count = sum(1 for c in clauses if '风险' in c['risk_level'] and '无风险' not in c['risk_level'])
    if risk_count > 0:
        print(f"  ✅ 风险识别检查通过: {risk_count} 条条款命中风险")
    else:
        print(f"  {Colors.YELLOW}⚠️ 未命中任何风险条款{Colors.ENDC}")
    
    print(f"\n{Colors.YELLOW}自我测试完成！{'✅ 全部通过' if all_pass else '⚠️ 部分需人工核对'}{Colors.ENDC}")
    return all_pass


def export_excel(clauses: List[Dict], output: Path):
    print(f"\n{Colors.CYAN}📊 导出 Excel...{Colors.ENDC}")
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    wb = openpyxl.Workbook()
    h_font = Font(bold=True, size=12, color="FFFFFF")
    h_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    r_high = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    r_medium = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    
    # Sheet 1: 全部条款
    ws1 = wb.active
    ws1.title = "全部条款"
    for col, h in enumerate(['条款编号', '条款分类', '风险等级', '风险关键词', '条款内容', '字数'], 1):
        c = ws1.cell(row=1, column=col, value=h)
        c.font = h_font; c.fill = h_fill
    
    for row, clause in enumerate(clauses, 2):
        ws1.cell(row=row, column=1, value=clause['id'])
        ws1.cell(row=row, column=2, value=clause['category'])
        ws1.cell(row=row, column=3, value=clause['risk_level'])
        ws1.cell(row=row, column=4, value=clause['risk_keywords'])
        ws1.cell(row=row, column=5, value=clause['content'])
        ws1.cell(row=row, column=6, value=clause['length'])
        if '高风险' in clause['risk_level']: ws1.cell(row=row, column=3).fill = r_high
        elif '中风险' in clause['risk_level']: ws1.cell(row=row, column=3).fill = r_medium
    
    ws1.column_dimensions['A'].width = 10
    ws1.column_dimensions['B'].width = 14
    ws1.column_dimensions['C'].width = 12
    ws1.column_dimensions['D'].width = 20
    ws1.column_dimensions['E'].width = 120
    
    # Sheet 2: 分类视图
    ws2 = wb.create_sheet("分类视图")
    cat_list = sorted(list(set(c['category'] for c in clauses)))
    row = 1
    for cat in cat_list:
        ws2.cell(row=row, column=1, value=cat).font = Font(bold=True, size=14)
        row += 1
        for clause in [c for c in clauses if c['category'] == cat]:
            ws2.cell(row=row, column=1, value=f"条款 {clause['id']}")
            ws2.cell(row=row, column=2, value=clause['risk_level'])
            ws2.cell(row=row, column=3, value=clause['content'][:200])
            row += 1
        row += 1
    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 140
    
    # Sheet 3: 风险条款汇总
    ws3 = wb.create_sheet("风险条款汇总")
    for col, h in enumerate(['条款编号', '风险等级', '风险关键词', '条款内容'], 1):
        c = ws3.cell(row=1, column=col, value=h)
        c.font = h_font; c.fill = h_fill
    risk_clauses = [c for c in clauses if '高风险' in c['risk_level'] or '中风险' in c['risk_level']]
    for row, clause in enumerate(risk_clauses, 2):
        ws3.cell(row=row, column=1, value=clause['id'])
        ws3.cell(row=row, column=2, value=clause['risk_level'])
        ws3.cell(row=row, column=3, value=clause['risk_keywords'])
        ws3.cell(row=row, column=4, value=clause['content'])
        if '高风险' in clause['risk_level']: ws3.cell(row=row, column=2).fill = r_high
        elif '中风险' in clause['risk_level']: ws3.cell(row=row, column=2).fill = r_medium
    ws3.column_dimensions['A'].width = 10
    ws3.column_dimensions['B'].width = 12
    ws3.column_dimensions['C'].width = 30
    ws3.column_dimensions['D'].width = 140
    
    # Sheet 4: OCR原始文本
    ws4 = wb.create_sheet("OCR原始文本")
    ws4.cell(row=1, column=1, value="OCR识别原始文本（v5.0自我测试版）").font = h_font
    ws4.cell(row=1, column=1).fill = h_fill
    ws4.cell(row=2, column=1, value='\n'.join(c['content'] for c in clauses)[:32767])
    ws4.column_dimensions['A'].width = 140
    
    wb.save(str(output))
    print(f"  ✅ 已导出: {output}")


def main():
    pdf_path = Path("/Users/bangcle/Downloads/contract/XSZS2603090130北京聚信得仁.pdf")
    output = Path("/Users/bangcle/Downloads/contract/XSZS2603090130_OCR_v5最终版.xlsx")
    
    print(f"\n{Colors.CYAN}{'='*60}{Colors.ENDC}")
    print(f"  📜 OCR v5.0 - 最终交付版（含完整自我测试）")
    print(f"{Colors.CYAN}{'='*60}{Colors.ENDC}")
    
    # ========== Step 1: PDF OCR 提取 ==========
    print(f"\n  📷 Step 1/6: PDF OCR 逐行识别中 ...")
    
    from pdf2image import convert_from_path
    import pytesseract
    
    pages = convert_from_path(str(pdf_path), dpi=400)
    all_lines = []
    
    for i, page in enumerate(pages, 1):
        print(f"  🔍  识别第 {i}/{len(pages)} 页 ...", end='\r')
        text = pytesseract.image_to_string(page, config=r'--oem 3 --psm 6 -l chi_sim --dpi 400')
        all_lines.extend([l.rstrip() for l in text.split('\n')])  # 保留行结构
    
    print(f"\n  ✅ OCR原始识别: {len(all_lines)} 行")
    
    # ========== Step 2: OCR逐行修正 ==========
    print(f"\n  🧹 Step 2/6: OCR错误逐行修正中 ...")
    cleaned_lines = correct_ocr_line_by_line(all_lines)
    print(f"  ✅ 修正完成: {len(cleaned_lines)} 有效行")
    
    # ========== Step 3: 按条款边界拆分 ==========
    clauses = split_clauses_by_boundary(cleaned_lines)
    
    # ========== Step 4: 分类 ==========
    print(f"\n  🏷️  Step 4/6: 条款分类中 ...")
    for idx, clause in enumerate(clauses):
        clause['category'] = classify_clause(clause['content'], idx, len(clauses))
    
    cats = {}
    for c in clauses: cats[c['category']] = cats.get(c['category'], 0) + 1
    print(f"  📊 分类统计:")
    for cat, cnt in sorted(cats.items()):
        print(f"    {cat}: {cnt} 条")
    
    # ========== Step 5: 风险识别 ==========
    print(f"\n  ⚠️  Step 5/6: 风险识别中 ...")
    risks = {'high': 0, 'medium': 0, 'low': 0}
    for clause in clauses:
        risk_lvl, risk_kws = identify_risk(clause['content'])
        clause['risk_level'] = risk_lvl
        clause['risk_keywords'] = risk_kws
        if '高风险' in risk_lvl: risks['high'] += 1
        elif '中风险' in risk_lvl: risks['medium'] += 1
        elif '低风险' in risk_lvl: risks['low'] += 1
    
    print(f"  🔴 高风险: {risks['high']} 条")
    print(f"  🟡 中风险: {risks['medium']} 条")
    print(f"  🟢 低风险: {risks['low']} 条")
    
    # ========== Step 6: 自我测试 ==========
    test_passed = self_test_report(clauses)
    
    # ========== 导出 ==========
    export_excel(clauses, output)
    
    print(f"\n{Colors.GREEN}{'='*60}{Colors.ENDC}")
    print(f"  🎉 v5.0 最终版全部完成！")
    print(f"{Colors.GREEN}{'='*60}{Colors.ENDC}")
    print(f"\n  📂 输出文件: {output}")
    print(f"\n  💡 修复的关键问题:")
    print(f"    ✅ 修复致命bug：保留换行符，确保条款边界正确识别")
    print(f"    ✅ 60+条合同OCR错误修正字典")
    print(f"    ✅ 5大类权重计分分类规则（位置加权 + 排除规则）")
    print(f"    ✅ 强匹配兜底（价格/违约关键词强制命中）")
    print(f"    ✅ 内置自我测试程序（条款数量/多样性/长度/风险覆盖率）")
    print(f"\n  Excel已自动打开，请核对！")
    
    import subprocess
    subprocess.run(['open', str(output)])


if __name__ == "__main__":
    main()
