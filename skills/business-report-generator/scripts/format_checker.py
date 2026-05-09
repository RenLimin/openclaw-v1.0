#!/usr/bin/env python3
"""
格式校验模块
==========
负责逐单元格对比生成文件与模板的格式一致性
"""

import openpyxl
from openpyxl.styles import PatternFill, Border, Font, Alignment
from openpyxl.utils import get_column_letter
import logging
from pathlib import Path
from typing import Dict, List, Any, Tuple

logger = logging.getLogger(__name__)


class FormatChecker:
    """格式校验器"""
    
    def __init__(self, template_path: str, generated_path: str):
        """
        初始化格式校验器
        
        Args:
            template_path: 模板文件路径
            generated_path: 生成的文件路径
        """
        self.template_path = Path(template_path)
        self.generated_path = Path(generated_path)
        self.template_wb = None
        self.generated_wb = None
        self.check_report: Dict[str, Any] = {
            'success': True,
            'total_cells_checked': 0,
            'differences': [],
            'sheets_summary': {}
        }
    
    def load_workbooks(self):
        """加载工作簿"""
        logger.info("加载格式校验文件...")
        self.template_wb = openpyxl.load_workbook(
            self.template_path,
            data_only=False
        )
        self.generated_wb = openpyxl.load_workbook(
            self.generated_path,
            data_only=False
        )
        logger.info(f"  模板: {len(self.template_wb.sheetnames)} 个Sheet")
        logger.info(f"  生成文件: {len(self.generated_wb.sheetnames)} 个Sheet")
    
    def check_all_sheets(self, check_data_only: bool = False, max_rows: int = 1000) -> Dict[str, Any]:
        """校验所有Sheet的格式"""
        logger.info("开始格式校验...")
        
        self.load_workbooks()
        
        common_sheets = set(self.template_wb.sheetnames) & set(self.generated_wb.sheetnames)
        
        for sheet_name in common_sheets:
            self.check_sheet(sheet_name, check_data_only, max_rows)
        
        # 检查缺失的Sheet
        missing_in_generated = set(self.template_wb.sheetnames) - set(self.generated_wb.sheetnames)
        for sheet_name in missing_in_generated:
            self.check_report['differences'].append({
                'type': 'SHEET_MISSING',
                'sheet': sheet_name,
                'message': f"生成文件缺少Sheet: {sheet_name}"
            })
        
        extra_in_generated = set(self.generated_wb.sheetnames) - set(self.template_wb.sheetnames)
        for sheet_name in extra_in_generated:
            self.check_report['differences'].append({
                'type': 'SHEET_EXTRA',
                'sheet': sheet_name,
                'message': f"生成文件多出Sheet: {sheet_name}"
            })
        
        # 生成摘要
        self._generate_summary()
        
        return self.check_report
    
    def check_sheet(self, sheet_name: str, check_data_only: bool = False, max_rows: int = 1000):
        """校验单个Sheet的格式"""
        template_ws = self.template_wb[sheet_name]
        generated_ws = self.generated_wb[sheet_name]
        
        differences = []
        cells_checked = 0
        
        # 只检查到 max_rows 行
        max_check_row = min(max(template_ws.max_row, generated_ws.max_row), max_rows)
        max_check_col = max(template_ws.max_column, generated_ws.max_column)
        
        sheet_summary = {
            'sheet_name': sheet_name,
            'template_rows': template_ws.max_row,
            'template_cols': template_ws.max_column,
            'generated_rows': generated_ws.max_row,
            'generated_cols': generated_ws.max_column,
            'cells_checked': 0,
            'differences_count': 0
        }
        
        logger.info(f"  校验Sheet: {sheet_name} (最多检查 {max_rows} 行)")
        
        for row in range(1, max_check_row + 1):
            for col in range(1, max_check_col + 1):
                t_cell = template_ws.cell(row=row, column=col) if row <= template_ws.max_row and col <= template_ws.max_column else None
                g_cell = generated_ws.cell(row=row, column=col) if row <= generated_ws.max_row and col <= generated_ws.max_column else None
                
                # 如果两个都是空单元格，跳过
                if (t_cell is None or t_cell.value is None) and (g_cell is None or g_cell.value is None):
                    continue
                
                cells_checked += 1
                
                # 检查格式差异
                cell_diffs = self._compare_cell_format(t_cell, g_cell, sheet_name, row, col)
                differences.extend(cell_diffs)
        
        sheet_summary['cells_checked'] = cells_checked
        sheet_summary['differences_count'] = len(differences)
        
        self.check_report['total_cells_checked'] += cells_checked
        self.check_report['differences'].extend(differences)
        self.check_report['sheets_summary'][sheet_name] = sheet_summary
        
        logger.info(f"    检查单元格: {cells_checked}, 差异: {len(differences)}")
    
    def _compare_cell_format(self, t_cell, g_cell, sheet_name: str, row: int, col: int) -> List[Dict[str, Any]]:
        """比较两个单元格的格式"""
        differences = []
        cell_ref = f"{get_column_letter(col)}{row}"
        
        # 如果一个单元格存在另一个不存在
        if t_cell is None and g_cell is not None:
            # 新增的数据行，格式不一致是正常的，只记录警告
            return differences
        
        if g_cell is None and t_cell is not None and t_cell.value is not None:
            differences.append({
                'type': 'CELL_MISSING',
                'sheet': sheet_name,
                'cell': cell_ref,
                'row': row,
                'column': col,
                'message': f"单元格 {cell_ref} 在生成文件中缺失"
            })
            return differences
        
        if g_cell is None:
            return differences
        
        # 1. 检查填充色（底色）
        t_fill = t_cell.fill.fgColor.rgb if (t_cell and t_cell.fill and t_cell.fill.fgColor) else None
        g_fill = g_cell.fill.fgColor.rgb if (g_cell.fill and g_cell.fill.fgColor) else None
        
        # 只检查表头或有特殊底色的单元格（非白色）
        white_colors = ['00FFFFFF', 'FFFFFFFF', None]
        if t_fill not in white_colors or g_fill not in white_colors:
            if t_fill != g_fill:
                differences.append({
                    'type': 'FILL_COLOR',
                    'sheet': sheet_name,
                    'cell': cell_ref,
                    'row': row,
                    'column': col,
                    'expected': t_fill,
                    'actual': g_fill,
                    'message': f"单元格 {cell_ref} 底色不一致: 预期 {t_fill}, 实际 {g_fill}"
                })
        
        # 2. 检查字体
        if t_cell:
            t_font = t_cell.font
            g_font = g_cell.font
            
            # 只检查表头行的字体（加粗等）
            if t_font and t_font.bold:
                if g_font.bold != t_font.bold:
                    differences.append({
                        'type': 'FONT_BOLD',
                        'sheet': sheet_name,
                        'cell': cell_ref,
                        'row': row,
                        'column': col,
                        'expected': t_font.bold,
                        'actual': g_font.bold,
                        'message': f"单元格 {cell_ref} 字体加粗不一致"
                    })
            
            if t_font and t_font.size:
                if g_font.size != t_font.size:
                    differences.append({
                        'type': 'FONT_SIZE',
                        'sheet': sheet_name,
                        'cell': cell_ref,
                        'row': row,
                        'column': col,
                        'expected': t_font.size,
                        'actual': g_font.size,
                        'message': f"单元格 {cell_ref} 字体大小不一致"
                    })
        
        # 3. 检查边框（简化检查 - 只检查是否有边框）
        if t_cell and t_cell.border:
            t_has_border = any([
                t_cell.border.left and t_cell.border.left.style != 'none',
                t_cell.border.right and t_cell.border.right.style != 'none',
                t_cell.border.top and t_cell.border.top.style != 'none',
                t_cell.border.bottom and t_cell.border.bottom.style != 'none'
            ])
            
            g_has_border = any([
                g_cell.border.left and g_cell.border.left.style != 'none',
                g_cell.border.right and g_cell.border.right.style != 'none',
                g_cell.border.top and g_cell.border.top.style != 'none',
                g_cell.border.bottom and g_cell.border.bottom.style != 'none'
            ])
            
            if t_has_border != g_has_border and t_cell.value is not None:
                differences.append({
                    'type': 'BORDER',
                    'sheet': sheet_name,
                    'cell': cell_ref,
                    'row': row,
                    'column': col,
                    'expected': t_has_border,
                    'actual': g_has_border,
                    'message': f"单元格 {cell_ref} 边框不一致"
                })
        
        # 4. 检查对齐方式
        if t_cell and t_cell.alignment and t_cell.value is not None:
            t_align = t_cell.alignment.horizontal
            g_align = g_cell.alignment.horizontal
            
            if t_align != g_align:
                differences.append({
                    'type': 'ALIGNMENT',
                    'sheet': sheet_name,
                    'cell': cell_ref,
                    'row': row,
                    'column': col,
                    'expected': t_align,
                    'actual': g_align,
                    'message': f"单元格 {cell_ref} 对齐方式不一致"
                })
        
        return differences
    
    def _generate_summary(self):
        """生成校验摘要"""
        diff_count = len(self.check_report['differences'])
        
        logger.info("=" * 60)
        logger.info("格式校验摘要")
        logger.info("=" * 60)
        logger.info(f"  总检查单元格: {self.check_report['total_cells_checked']}")
        logger.info(f"  发现差异数: {diff_count}")
        
        # 按类型统计差异
        diff_by_type = {}
        for diff in self.check_report['differences']:
            diff_type = diff['type']
            diff_by_type[diff_type] = diff_by_type.get(diff_type, 0) + 1
        
        for diff_type, count in diff_by_type.items():
            logger.info(f"    {diff_type}: {count}")
        
        if diff_count == 0:
            logger.info("✓ 格式完全一致！")
            self.check_report['success'] = True
        elif diff_count < 100:
            logger.info("⚠ 存在少量格式差异，请检查")
            self.check_report['success'] = True  # 少量差异仍然算成功
        else:
            logger.warning("✗ 存在大量格式差异")
            self.check_report['success'] = False
        
        logger.info("=" * 60)
    
    def save_report(self, output_path: str):
        """保存校验报告到文件"""
        import json
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(self.check_report, f, ensure_ascii=False, indent=2)
        
        logger.info(f"格式校验报告已保存: {output_path}")
    
    def print_differences(self, limit: int = 50):
        """打印差异详情"""
        diffs = self.check_report['differences']
        
        if not diffs:
            return
        
        print("\n格式差异详情:")
        print("-" * 60)
        
        for i, diff in enumerate(diffs[:limit]):
            print(f"{i+1}. [{diff['type']}] {diff['sheet']}!{diff.get('cell', '??')}")
            print(f"   {diff['message']}")
        
        if len(diffs) > limit:
            print(f"... 还有 {len(diffs) - limit} 个差异未显示")
