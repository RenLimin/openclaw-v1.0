#!/usr/bin/env python3
"""
公式保护模块
============
智能检测公式列，写入时自动跳过，支持公式自动计算刷新，
公式完整性校验（写入前后一致性检查）
"""

import logging
from pathlib import Path
from typing import Dict, List, Any, Set, Tuple
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import yaml

logger = logging.getLogger(__name__)


class FormulaProtector:
    """
    公式保护与自动计算类
    
    功能:
    - 智能检测公式列
    - 写入时自动跳过公式列
    - 公式完整性校验
    - 标记需要刷新的公式
    """
    
    def __init__(self, workbook_path: str, config_path: str = None, data_only: bool = False):
        """
        加载工作簿
        
        Args:
            workbook_path: 工作簿路径
            config_path: 配置文件路径
            data_only: 是否只读取值（不读取公式）
        """
        self.workbook_path = Path(workbook_path)
        self.config = self._load_config(config_path)
        
        if not self.workbook_path.exists():
            raise FileNotFoundError(f"工作簿不存在: {workbook_path}")
        
        self.workbook = openpyxl.load_workbook(
            self.workbook_path,
            data_only=data_only,
            keep_vba=False
        )
        
        # 缓存检测到的公式列
        self._formula_columns_cache: Dict[str, List[int]] = {}
        
        # 缓存原始公式（用于校验）
        self._original_formulas_cache: Dict[str, Dict[str, str]] = {}
        
        logger.info(f"已加载工作簿: {self.workbook_path.name}")
        logger.info(f"  Sheet数量: {len(self.workbook.sheetnames)}")
    
    def _load_config(self, config_path: str = None) -> Dict[str, Any]:
        """加载配置文件"""
        if config_path is None:
            config_path = Path(__file__).parent.parent / 'config' / 'formula-rules.yaml'
        
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                return yaml.safe_load(f) or {}
        except Exception as e:
            logger.warning(f"加载配置文件失败，使用默认配置: {e}")
            return {}
    
    def detect_formula_columns(self, sheet_name: str, analyze_rows: int = None) -> List[int]:
        """
        检测公式列
        
        Args:
            sheet_name: Sheet名称
            analyze_rows: 分析的行数，默认使用配置值
            
        Returns:
            公式列的列号列表（从1开始）
        """
        if sheet_name in self._formula_columns_cache:
            return self._formula_columns_cache[sheet_name]
        
        detection_config = self.config.get('detection', {})
        if analyze_rows is None:
            analyze_rows = detection_config.get('analyze_rows', 50)
        
        threshold = detection_config.get('formula_column_threshold', 3)
        formula_prefix = detection_config.get('formula_prefix', '=')
        
        if sheet_name not in self.workbook.sheetnames:
            logger.warning(f"Sheet不存在: {sheet_name}")
            return []
        
        ws = self.workbook[sheet_name]
        
        # 找出有公式的列
        formula_counts: Dict[int, int] = {}
        formulas_by_cell: Dict[str, str] = {}
        
        for row in range(1, min(ws.max_row + 1, analyze_rows + 1)):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell_ref = f"{get_column_letter(col)}{row}"
                
                if cell.value and isinstance(cell.value, str) and cell.value.startswith(formula_prefix):
                    formula_counts[col] = formula_counts.get(col, 0) + 1
                    formulas_by_cell[cell_ref] = cell.value
        
        # 缓存原始公式
        self._original_formulas_cache[sheet_name] = formulas_by_cell
        
        # 识别公式列（连续公式数超过阈值）
        formula_columns: List[int] = [
            col for col, count in formula_counts.items()
            if count >= threshold
        ]
        
        self._formula_columns_cache[sheet_name] = formula_columns
        
        logger.info(f"Sheet '{sheet_name}' 检测到 {len(formula_columns)} 个公式列")
        logger.debug(f"  公式列: {[get_column_letter(c) for c in formula_columns]}")
        
        return formula_columns
    
    def get_formulas_in_sheet(self, sheet_name: str) -> Dict[str, str]:
        """
        获取Sheet中所有公式
        
        Args:
            sheet_name: Sheet名称
            
        Returns:
            {cell_reference: formula_string} 字典
        """
        if sheet_name not in self._original_formulas_cache:
            self.detect_formula_columns(sheet_name)
        
        return self._original_formulas_cache.get(sheet_name, {})
    
    def write_data(self, sheet_name: str, data: pd.DataFrame, 
                   start_row: int = 2, skip_formula_columns: bool = None):
        """
        写入数据，自动跳过公式列
        
        Args:
            sheet_name: Sheet名称
            data: 要写入的数据
            start_row: 起始行号（从1开始）
            skip_formula_columns: 是否跳过公式列，默认使用配置值
        """
        protection_config = self.config.get('protection', {})
        if skip_formula_columns is None:
            skip_formula_columns = protection_config.get('skip_formula_columns', True)
        
        if sheet_name not in self.workbook.sheetnames:
            ws = self.workbook.create_sheet(sheet_name)
            logger.info(f"创建新Sheet: {sheet_name}")
        else:
            ws = self.workbook[sheet_name]
        
        # 获取公式列
        formula_columns = set()
        if skip_formula_columns:
            formula_columns = set(self.detect_formula_columns(sheet_name))
        
        logger.info(f"写入数据到Sheet '{sheet_name}': {len(data)} 行 x {len(data.columns)} 列")
        if formula_columns:
            logger.info(f"  跳过公式列: {[get_column_letter(c) for c in formula_columns]}")
        
        # 写入数据
        rows_written = 0
        for df_idx, row_data in enumerate(data.itertuples(index=False)):
            excel_row = start_row + df_idx
            
            for col_idx, value in enumerate(row_data, start=1):
                # 跳过公式列
                if col_idx in formula_columns:
                    continue
                
                if pd.isna(value):
                    continue
                
                cell = ws.cell(row=excel_row, column=col_idx)
                cell.value = value
            
            rows_written += 1
            
            if rows_written % 10000 == 0:
                logger.info(f"  已写入 {rows_written} 行...")
        
        logger.info(f"  完成写入: {rows_written} 行")
    
    def write_data_column_aligned(self, sheet_name: str, data: pd.DataFrame,
                                  header_row: int = 1, skip_formula_columns: bool = None):
        """
        按列名对齐写入数据（根据表头匹配列位置）
        
        Args:
            sheet_name: Sheet名称
            data: 要写入的数据
            header_row: 表头行号
            skip_formula_columns: 是否跳过公式列
        """
        if sheet_name not in self.workbook.sheetnames:
            logger.warning(f"Sheet不存在: {sheet_name}")
            return
        
        ws = self.workbook[sheet_name]
        
        # 读取表头建立列映射
        column_mapping: Dict[str, int] = {}
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=header_row, column=col).value
            if cell_value:
                column_mapping[str(cell_value).strip()] = col
        
        # 获取公式列
        formula_columns = set()
        if skip_formula_columns is None or skip_formula_columns:
            formula_columns = set(self.detect_formula_columns(sheet_name))
        
        logger.info(f"按列名对齐写入: {len(data.columns)} 列匹配到 {len(column_mapping)} 个表头")
        
        # 写入数据
        start_row = header_row + 1
        for df_idx, row_data in data.iterrows():
            excel_row = start_row + df_idx
            
            for col_name, value in row_data.items():
                col_idx = column_mapping.get(str(col_name).strip())
                if col_idx is None:
                    continue
                
                if col_idx in formula_columns:
                    continue
                
                if pd.isna(value):
                    continue
                
                cell = ws.cell(row=excel_row, column=col_idx)
                cell.value = value
        
        logger.info(f"  完成写入: {len(data)} 行")
    
    def refresh_formulas(self, sheet_name: str = None):
        """
        刷新公式（标记需要Excel重新计算）
        
        注意: openpyxl 不实际计算公式，只是标记属性让Excel在打开时重新计算
        
        Args:
            sheet_name: Sheet名称，None表示全部Sheet
        """
        sheet_names = [sheet_name] if sheet_name else self.workbook.sheetnames
        
        for name in sheet_names:
            if name not in self.workbook.sheetnames:
                continue
            
            ws = self.workbook[name]
            # 标记工作簿需要重新计算
            ws.calculate_dimension()
            
            # 强制所有公式单元格重新计算
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        # 触发公式重新评估（实际上不会计算，但会标记）
                        cell.value = cell.value
        
        # 设置工作簿属性，强制Excel打开时完全重新计算
        self.workbook.calculation.fullCalcOnLoad = True
        
        logger.info(f"已标记公式需要重新计算")
    
    def verify_formula_integrity(self, sheet_name: str = None) -> Dict[str, Any]:
        """
        校验公式完整性
        
        Args:
            sheet_name: Sheet名称，None表示全部Sheet
            
        Returns:
            校验报告字典
        """
        validation_config = self.config.get('validation', {})
        report_level = validation_config.get('report_level', 'detailed')
        
        report = {
            'success': True,
            'total_formulas': 0,
            'changed_formulas': 0,
            'lost_formulas': 0,
            'details': []
        }
        
        sheet_names = [sheet_name] if sheet_name else self.workbook.sheetnames
        
        for name in sheet_names:
            if name not in self.workbook.sheetnames:
                continue
            
            if name not in self._original_formulas_cache:
                logger.warning(f"未缓存Sheet '{name}' 的原始公式，跳过校验")
                continue
            
            original_formulas = self._original_formulas_cache[name]
            ws = self.workbook[name]
            
            report['total_formulas'] += len(original_formulas)
            
            for cell_ref, original_formula in original_formulas.items():
                current_value = ws[cell_ref].value
                
                # 检查公式是否丢失
                if not current_value or not isinstance(current_value, str):
                    report['lost_formulas'] += 1
                    report['success'] = False
                    if report_level == 'detailed':
                        report['details'].append({
                            'type': 'FORMULA_LOST',
                            'sheet': name,
                            'cell': cell_ref,
                            'message': f"公式丢失: {cell_ref}"
                        })
                    continue
                
                # 检查公式是否变更
                if current_value != original_formula:
                    # 检查是否是相对引用的正常变化
                    if self._is_relative_reference_change(original_formula, current_value):
                        continue  # 正常变化，忽略
                    
                    report['changed_formulas'] += 1
                    if report_level == 'detailed':
                        report['details'].append({
                            'type': 'FORMULA_CHANGED',
                            'sheet': name,
                            'cell': cell_ref,
                            'expected': original_formula,
                            'actual': current_value,
                            'message': f"公式变更: {cell_ref}"
                        })
        
        logger.info(f"公式完整性校验: 总数 {report['total_formulas']}, "
                   f"变更 {report['changed_formulas']}, 丢失 {report['lost_formulas']}")
        
        return report
    
    def _is_relative_reference_change(self, formula1: str, formula2: str) -> bool:
        """检查是否是相对引用的正常变化"""
        # 简单检查：去除数字后是否相同
        import re
        stripped1 = re.sub(r'\d+', '', formula1)
        stripped2 = re.sub(r'\d+', '', formula2)
        return stripped1 == stripped2
    
    def restore_formulas(self, sheet_name: str = None):
        """
        恢复公式到原始状态（如果被意外覆盖）
        
        Args:
            sheet_name: Sheet名称，None表示全部Sheet
        """
        sheet_names = [sheet_name] if sheet_name else self.workbook.sheetnames
        
        restored_count = 0
        for name in sheet_names:
            if name not in self._original_formulas_cache:
                continue
            
            original_formulas = self._original_formulas_cache[name]
            ws = self.workbook[name]
            
            for cell_ref, original_formula in original_formulas.items():
                current_value = ws[cell_ref].value
                if current_value != original_formula:
                    ws[cell_ref].value = original_formula
                    restored_count += 1
        
        logger.info(f"已恢复 {restored_count} 个公式")
    
    def save(self, output_path: str):
        """
        保存工作簿
        
        Args:
            output_path: 输出文件路径
        """
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # 保存前校验（如果配置要求）
        validation_config = self.config.get('validation', {})
        if validation_config.get('verify_after_write', True):
            self.verify_formula_integrity()
        
        # 标记需要刷新
        refresh_config = self.config.get('auto_refresh', {})
        if refresh_config.get('mark_for_refresh', True):
            self.refresh_formulas()
        
        self.workbook.save(str(output_path))
        logger.info(f"工作簿已保存: {output_path}")
    
    def close(self):
        """关闭工作簿"""
        self.workbook.close()
    
    def __enter__(self):
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()
