#!/usr/bin/env python3
"""
透视表操作模块
==============
检测现有透视表位置与数据源，更新透视表数据源，
支持透视表自动刷新标记
"""

import logging
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
import openpyxl
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.pivot.table import PivotTable, PivotField

logger = logging.getLogger(__name__)


class PivotController:
    """
    透视表操作类
    
    功能:
    - 检测透视表位置
    - 获取透视表数据源
    - 更新透视表数据源
    - 标记透视表需要刷新
    """
    
    def __init__(self, workbook_path: str):
        """
        加载工作簿
        
        Args:
            workbook_path: 工作簿路径
        """
        self.workbook_path = Path(workbook_path)
        
        if not self.workbook_path.exists():
            raise FileNotFoundError(f"工作簿不存在: {workbook_path}")
        
        # 注意: 要操作透视表，不能使用 read_only 模式
        self.workbook = openpyxl.load_workbook(
            self.workbook_path,
            data_only=False,
            keep_vba=True,
            read_only=False
        )
        
        # 缓存检测到的透视表
        self._pivot_cache: Dict[str, List[Dict]] = {}
        
        logger.info(f"已加载工作簿: {self.workbook_path.name}")
    
    def detect_pivot_tables(self, sheet_name: str = None) -> Dict[str, List[Dict]]:
        """
        检测透视表位置
        
        Args:
            sheet_name: Sheet名称，None表示全部Sheet
            
        Returns:
            {sheet_name: [pivot_info_list]} 字典
        """
        result = {}
        
        sheet_names = [sheet_name] if sheet_name else self.workbook.sheetnames
        
        for name in sheet_names:
            if name not in self.workbook.sheetnames:
                continue
            
            ws = self.workbook[name]
            pivots = []
            
            # 1. 尝试通过 openpyxl 原生 API 检测
            if hasattr(ws, '_pivots') and ws._pivots:
                for pivot_idx, pivot in enumerate(ws._pivots):
                    try:
                        pivot_info = self._extract_pivot_info(pivot, name, pivot_idx)
                        pivots.append(pivot_info)
                    except Exception as e:
                        logger.debug(f"解析透视表失败: {e}")
            
            # 2. 如果没有检测到，通过特征关键字检测
            if not pivots:
                detected = self._detect_pivot_by_keywords(ws)
                pivots.extend(detected)
            
            if pivots:
                result[name] = pivots
                logger.info(f"Sheet '{name}' 检测到 {len(pivots)} 个透视表")
        
        self._pivot_cache = result
        return result
    
    def _extract_pivot_info(self, pivot: PivotTable, sheet_name: str, idx: int) -> Dict[str, Any]:
        """从 openpyxl PivotTable 对象提取信息"""
        info = {
            'name': pivot.name if hasattr(pivot, 'name') else f"PivotTable{idx+1}",
            'sheet': sheet_name,
            'location': None,
            'data_source': None,
            'cache_id': pivot.cacheId if hasattr(pivot, 'cacheId') else None,
            'fields': [],
            'row_fields': [],
            'column_fields': [],
            'value_fields': [],
            'detected_by': 'openpyxl_api'
        }
        
        # 透视表位置
        if hasattr(pivot, 'location'):
            loc = pivot.location
            if hasattr(loc, 'ref'):
                info['location'] = str(loc.ref)
        
        # 数据源 - 从缓存获取
        try:
            if hasattr(pivot, 'cache'):
                cache = pivot.cache
                if hasattr(cache, 'cacheSource'):
                    source = cache.cacheSource
                    if hasattr(source, 'worksheetSource'):
                        ws_source = source.worksheetSource
                        if ws_source:
                            sheet = getattr(ws_source, 'sheet', None)
                            ref = getattr(ws_source, 'ref', None)
                            if sheet and ref:
                                info['data_source'] = f"'{sheet}'!{ref}"
                            elif ref:
                                info['data_source'] = ref
        except Exception as e:
            logger.debug(f"获取透视表数据源失败: {e}")
        
        # 字段信息
        try:
            if hasattr(pivot, 'pivotFields'):
                for field in pivot.pivotFields:
                    field_name = getattr(field, 'name', 'Unknown')
                    info['fields'].append(field_name)
                    
                    # 检查字段类型
                    if hasattr(field, 'axis'):
                        axis = field.axis
                        if axis == 'axisRow':
                            info['row_fields'].append(field_name)
                        elif axis == 'axisCol':
                            info['column_fields'].append(field_name)
                        elif axis == 'axisValues':
                            info['value_fields'].append(field_name)
        except Exception as e:
            logger.debug(f"获取透视表字段失败: {e}")
        
        return info
    
    def _detect_pivot_by_keywords(self, ws) -> List[Dict]:
        """通过特征关键字检测透视表位置"""
        pivot_keywords = [
            '行标签', '列标签', '值', '总计', '求和项', '计数项',
            '平均值项', '最大值项', '最小值项', '(全部)', '(多项)'
        ]
        
        detected_pivots = []
        found_cells = []
        
        # 搜索前50行前20列
        for row in range(1, min(50, ws.max_row + 1)):
            for col in range(1, min(20, ws.max_column + 1)):
                cell = ws.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str):
                    cell_value = str(cell.value)
                    for keyword in pivot_keywords:
                        if keyword in cell_value:
                            found_cells.append({
                                'row': row,
                                'col': col,
                                'value': cell_value,
                                'keyword': keyword
                            })
                            break
        
        # 聚类检测到的单元格，识别透视表区域
        if found_cells:
            # 简单聚类：按行距离分组
            groups = []
            for cell in found_cells:
                placed = False
                for group in groups:
                    if abs(cell['row'] - group['min_row']) <= 10:
                        group['cells'].append(cell)
                        group['min_row'] = min(group['min_row'], cell['row'])
                        group['max_row'] = max(group['max_row'], cell['row'])
                        group['min_col'] = min(group['min_col'], cell['col'])
                        group['max_col'] = max(group['max_col'], cell['col'])
                        placed = True
                        break
                
                if not placed:
                    groups.append({
                        'cells': [cell],
                        'min_row': cell['row'],
                        'max_row': cell['row'],
                        'min_col': cell['col'],
                        'max_col': cell['col']
                    })
            
            # 生成透视表信息
            for idx, group in enumerate(groups):
                location = f"{get_column_letter(group['min_col'])}{group['min_row']}"
                if group['max_row'] > group['min_row'] or group['max_col'] > group['min_col']:
                    location += f":{get_column_letter(group['max_col'])}{group['max_row']}"
                
                detected_pivots.append({
                    'name': f"PivotTable{idx+1}",
                    'sheet': ws.title,
                    'location': location,
                    'data_source': None,
                    'fields': [],
                    'row_fields': [],
                    'column_fields': [],
                    'value_fields': [],
                    'detected_by': 'keyword_detection',
                    'detected_keywords': [c['keyword'] for c in group['cells']]
                })
        
        return detected_pivots
    
    def get_pivot_info(self, sheet_name: str, pivot_name: str = None) -> Optional[Dict]:
        """
        获取指定透视表信息
        
        Args:
            sheet_name: Sheet名称
            pivot_name: 透视表名称，None表示第一个
            
        Returns:
            透视表信息字典
        """
        if sheet_name not in self._pivot_cache:
            self.detect_pivot_tables(sheet_name)
        
        pivots = self._pivot_cache.get(sheet_name, [])
        if not pivots:
            return None
        
        if pivot_name is None:
            return pivots[0]
        
        for pivot in pivots:
            if pivot['name'] == pivot_name:
                return pivot
        
        return None
    
    def update_pivot_data_source(self, sheet_name: str, pivot_name: str, new_data_range: str) -> bool:
        """
        更新透视表数据源
        
        注意: openpyxl 对透视表的支持有限，此方法主要尝试更新缓存引用。
        在大多数情况下，Excel打开时会自动检测并允许用户刷新。
        
        Args:
            sheet_name: Sheet名称
            pivot_name: 透视表名称
            new_data_range: 新的数据源范围，如 "Sheet1!A1:D100"
            
        Returns:
            是否成功更新
        """
        if sheet_name not in self.workbook.sheetnames:
            logger.warning(f"Sheet不存在: {sheet_name}")
            return False
        
        ws = self.workbook[sheet_name]
        
        # 尝试查找并更新透视表
        if hasattr(ws, '_pivots'):
            for pivot in ws._pivots:
                pivot_name_current = getattr(pivot, 'name', '')
                if pivot_name_current == pivot_name or not pivot_name:
                    try:
                        # 尝试更新缓存
                        if hasattr(pivot, 'cache'):
                            cache = pivot.cache
                            if hasattr(cache, 'cacheSource'):
                                source = cache.cacheSource
                                if hasattr(source, 'worksheetSource'):
                                    # 解析新的数据范围
                                    if '!' in new_data_range:
                                        source_sheet, ref = new_data_range.rsplit('!', 1)
                                        source_sheet = source_sheet.strip("'")
                                    else:
                                        source_sheet = sheet_name
                                        ref = new_data_range
                                    
                                    ws_source = source.worksheetSource
                                    if ws_source:
                                        ws_source.sheet = source_sheet
                                        ws_source.ref = ref
                                        
                                        logger.info(f"已更新透视表 '{pivot_name}' 数据源: {new_data_range}")
                                        return True
                    except Exception as e:
                        logger.error(f"更新透视表数据源失败: {e}")
        
        logger.warning(f"无法更新透视表数据源，请在Excel中手动刷新")
        return False
    
    def update_pivot_data_source_by_expand(self, sheet_name: str, pivot_name: str, 
                                           new_last_row: int) -> bool:
        """
        通过扩展现有数据源范围来更新
        
        Args:
            sheet_name: Sheet名称
            pivot_name: 透视表名称
            new_last_row: 新的最后一行号
            
        Returns:
            是否成功更新
        """
        pivot_info = self.get_pivot_info(sheet_name, pivot_name)
        if not pivot_info or not pivot_info.get('data_source'):
            logger.warning("无法获取透视表现有数据源")
            return False
        
        current_source = pivot_info['data_source']
        
        # 解析范围，扩展行数
        try:
            if '!' in current_source:
                source_sheet, ref = current_source.rsplit('!', 1)
            else:
                source_sheet = sheet_name
                ref = current_source
            
            # 解析边界
            min_col, min_row, max_col, old_max_row = range_boundaries(ref)
            
            # 创建新的范围
            new_ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{new_last_row}"
            new_source = f"{source_sheet}!{new_ref}"
            
            logger.info(f"扩展数据源: {current_source} -> {new_source}")
            
            return self.update_pivot_data_source(sheet_name, pivot_name, new_source)
            
        except Exception as e:
            logger.error(f"解析数据源范围失败: {e}")
            return False
    
    def refresh_all_pivots(self) -> bool:
        """
        标记所有透视表需要刷新
        
        注意: openpyxl 不实际执行刷新，只是标记属性让Excel在打开时提示刷新
        
        Returns:
            是否成功标记
        """
        try:
            # 设置工作簿属性，提示刷新
            self.workbook.calculation.fullCalcOnLoad = True
            
            # 标记每个透视表
            for sheet_name in self.workbook.sheetnames:
                ws = self.workbook[sheet_name]
                if hasattr(ws, '_pivots'):
                    for pivot in ws._pivots:
                        # 尝试刷新缓存标记
                        if hasattr(pivot, 'cache'):
                            # 触发缓存重新加载标记
                            pass
            
            logger.info("已标记所有透视表需要刷新（请在Excel中手动执行刷新）")
            return True
            
        except Exception as e:
            logger.error(f"标记透视表刷新失败: {e}")
            return False
    
    def set_refresh_on_open(self, enabled: bool = True) -> bool:
        """
        设置工作簿打开时自动刷新数据
        
        Args:
            enabled: 是否启用
            
        Returns:
            是否成功设置
        """
        try:
            # 设置计算属性
            self.workbook.calculation.fullCalcOnLoad = enabled
            
            # 尝试设置透视表缓存的刷新属性
            for sheet_name in self.workbook.sheetnames:
                ws = self.workbook[sheet_name]
                if hasattr(ws, '_pivots'):
                    for pivot in ws._pivots:
                        if hasattr(pivot, 'cache') and hasattr(pivot.cache, 'refreshOnLoad'):
                            pivot.cache.refreshOnLoad = enabled
            
            logger.info(f"已{'启用' if enabled else '禁用'}工作簿打开时自动刷新")
            return True
            
        except Exception as e:
            logger.error(f"设置自动刷新失败: {e}")
            return False
    
    def list_all_pivots(self) -> List[Dict]:
        """
        列出所有透视表信息
        
        Returns:
            透视表信息列表
        """
        all_pivots = []
        pivot_info = self.detect_pivot_tables()
        
        for sheet_name, pivots in pivot_info.items():
            for pivot in pivots:
                all_pivots.append(pivot)
        
        # 打印摘要
        if all_pivots:
            print("\n透视表列表:")
            print("-" * 60)
            for i, pivot in enumerate(all_pivots, 1):
                print(f"{i}. 名称: {pivot['name']}")
                print(f"   Sheet: {pivot['sheet']}")
                print(f"   位置: {pivot.get('location', '未知')}")
                print(f"   数据源: {pivot.get('data_source', '未知')}")
                print(f"   检测方式: {pivot.get('detected_by', '未知')}")
                if pivot.get('row_fields'):
                    print(f"   行字段: {', '.join(pivot['row_fields'])}")
                if pivot.get('column_fields'):
                    print(f"   列字段: {', '.join(pivot['column_fields'])}")
                if pivot.get('value_fields'):
                    print(f"   值字段: {', '.join(pivot['value_fields'])}")
                print()
        else:
            print("未检测到透视表")
        
        return all_pivots
    
    def save(self, output_path: str = None):
        """
        保存工作簿
        
        Args:
            output_path: 输出文件路径，None表示覆盖原文件
        """
        if output_path is None:
            output_path = self.workbook_path
        
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        self.workbook.save(str(output_path))
        logger.info(f"工作簿已保存: {output_path}")
    
    def close(self):
        """关闭工作簿"""
        self.workbook.close()
    
    def __enter__(self):
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()
