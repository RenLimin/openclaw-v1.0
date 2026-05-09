#!/usr/bin/env python3
"""
模板分析模块
==========
负责分析Excel模板结构，识别公式、格式、透视表等
"""

import openpyxl
from openpyxl.styles import PatternFill, Border, Font, Alignment
from openpyxl.utils import get_column_letter
import logging
from pathlib import Path
from typing import Dict, List, Any, Tuple, Set

logger = logging.getLogger(__name__)


class TemplateAnalyzer:
    """模板分析器"""
    
    def __init__(self, template_path: str):
        """
        初始化模板分析器
        
        Args:
            template_path: 模板文件路径
        """
        self.template_path = Path(template_path)
        self.workbook = None
        self.template_info: Dict[str, Any] = {
            'sheets': {},
            'formula_columns': {},
            'formats': {},
            'pivot_tables': {},
            'data_ranges': {}
        }
        self.load_workbook()
    
    def load_workbook(self):
        """加载工作簿"""
        logger.info(f"加载模板: {self.template_path}")
        self.workbook = openpyxl.load_workbook(
            self.template_path,
            data_only=False,
            keep_vba=False
        )
        logger.info(f"  Sheet数量: {len(self.workbook.sheetnames)}")
    
    def analyze_all_sheets(self) -> Dict[str, Any]:
        """分析所有Sheet"""
        logger.info("开始分析模板结构...")
        
        for sheet_name in self.workbook.sheetnames:
            self.analyze_sheet(sheet_name)
        
        self._analyze_pivot_tables()
        
        logger.info("✓ 模板分析完成")
        return self.template_info
    
    def analyze_sheet(self, sheet_name: str):
        """分析单个Sheet"""
        ws = self.workbook[sheet_name]
        
        sheet_info = {
            'name': sheet_name,
            'max_row': ws.max_row,
            'max_column': ws.max_column,
            'dimensions': ws.dimensions,
            'headers': [],
            'formula_cells': [],
            'data_range': None,
            'header_row': None
        }
        
        # 查找表头行（第一行有值的行）
        header_row = None
        for row in range(1, min(10, ws.max_row + 1)):
            has_value = False
            headers = []
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None and str(cell.value).strip():
                    has_value = True
                    headers.append({
                        'column': col,
                        'column_letter': get_column_letter(col),
                        'header': str(cell.value).strip()
                    })
            if has_value and len(headers) > 3:
                header_row = row
                sheet_info['header_row'] = header_row
                sheet_info['headers'] = headers
                break
        
        # 识别公式单元格
        formula_cells = []
        for row in range(1, min(ws.max_row + 1, 100)):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula_cells.append({
                        'row': row,
                        'column': col,
                        'cell': f"{get_column_letter(col)}{row}",
                        'formula': cell.value
                    })
        
        sheet_info['formula_cells'] = formula_cells
        
        # 识别公式列（整列都是公式的列）
        formula_columns: Set[int] = set()
        for formula_cell in formula_cells:
            col = formula_cell['column']
            # 检查该列后续行是否也有相同模式的公式
            formula_count = 0
            for row in range(header_row + 1 if header_row else 2, min(ws.max_row + 1, 50)):
                cell = ws.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula_count += 1
            if formula_count >= 3:
                formula_columns.add(col)
        
        sheet_info['formula_columns'] = sorted(list(formula_columns))
        
        # 识别数据范围（从表头下一行开始到最后一行有数据的区域）
        if header_row and ws.max_row > header_row:
            sheet_info['data_range'] = {
                'start_row': header_row + 1,
                'end_row': ws.max_row,
                'start_column': 1,
                'end_column': ws.max_column
            }
        
        # 记录单元格格式（表头和前几行数据的格式）
        self._record_cell_formats(sheet_name, ws, header_row)
        
        self.template_info['sheets'][sheet_name] = sheet_info
        
        logger.info(f"  Sheet: {sheet_name}")
        logger.info(f"    大小: {ws.max_row}行 x {ws.max_column}列")
        logger.info(f"    表头行: {header_row}")
        logger.info(f"    公式单元格: {len(formula_cells)}")
        logger.info(f"    公式列: {len(formula_columns)}")
    
    def _record_cell_formats(self, sheet_name: str, ws, header_row: int = None):
        """记录单元格格式"""
        formats = {}
        
        # 记录表头格式
        if header_row:
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=header_row, column=col)
                cell_key = f"header_{col}"
                formats[cell_key] = self._extract_cell_format(cell)
        
        # 记录前几行数据的格式
        start_row = header_row + 1 if header_row else 1
        for row in range(start_row, min(start_row + 5, ws.max_row + 1)):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None:
                    cell_key = f"data_{row}_{col}"
                    formats[cell_key] = self._extract_cell_format(cell)
        
        self.template_info['formats'][sheet_name] = formats
    
    def _extract_cell_format(self, cell) -> Dict[str, Any]:
        """提取单元格格式信息"""
        return {
            'fill': {
                'fgColor': cell.fill.fgColor.rgb if cell.fill.fgColor else None,
                'patternType': cell.fill.patternType
            } if cell.fill else None,
            'font': {
                'name': cell.font.name,
                'size': cell.font.size,
                'bold': cell.font.bold,
                'italic': cell.font.italic,
                'color': cell.font.color.rgb if cell.font.color else None
            } if cell.font else None,
            'border': {
                'left': cell.border.left.style if cell.border.left else None,
                'right': cell.border.right.style if cell.border.right else None,
                'top': cell.border.top.style if cell.border.top else None,
                'bottom': cell.border.bottom.style if cell.border.bottom else None
            } if cell.border else None,
            'alignment': {
                'horizontal': cell.alignment.horizontal,
                'vertical': cell.alignment.vertical,
                'wrap_text': cell.alignment.wrap_text
            } if cell.alignment else None,
            'number_format': cell.number_format
        }
    
    def _analyze_pivot_tables(self):
        """分析透视表"""
        pivot_info = {}
        
        for sheet_name in self.workbook.sheetnames:
            ws = self.workbook[sheet_name]
            
            # openpyxl 的透视表支持有限，我们通过查找特殊的透视表区域来识别
            # 检查是否有透视表缓存
            sheet_pivots = []
            
            if hasattr(ws, '_pivots') and ws._pivots:
                for pivot in ws._pivots:
                    try:
                        pivot_info_item = {
                            'name': pivot.name if hasattr(pivot, 'name') else 'Unknown',
                            'location': pivot.location if hasattr(pivot, 'location') else None,
                            'cache_id': pivot.cacheId if hasattr(pivot, 'cacheId') else None
                        }
                        sheet_pivots.append(pivot_info_item)
                    except:
                        pass
            
            # 也可以通过特征识别：透视表通常有"行标签"、"列标签"、"值"、"总计"等关键词
            pivot_keywords = ['行标签', '列标签', '值', '总计', '求和项', '计数项', '(全部)', '(多项)']
            for row in range(1, min(20, ws.max_row + 1)):
                for col in range(1, min(20, ws.max_column + 1)):
                    cell = ws.cell(row=row, column=col)
                    if cell.value and isinstance(cell.value, str):
                        for keyword in pivot_keywords:
                            if keyword in cell.value:
                                # 可能是透视表区域
                                found = False
                                for p in sheet_pivots:
                                    if p.get('detected_row') == row:
                                        found = True
                                        break
                                if not found:
                                    sheet_pivots.append({
                                        'detected': True,
                                        'detected_row': row,
                                        'detected_col': col,
                                        'keyword': keyword
                                    })
                                break
            
            if sheet_pivots:
                pivot_info[sheet_name] = sheet_pivots
                logger.info(f"    透视表: {len(sheet_pivots)} 个")
        
        self.template_info['pivot_tables'] = pivot_info
    
    def get_sheet_info(self, sheet_name: str) -> Dict[str, Any]:
        """获取指定Sheet的信息"""
        return self.template_info['sheets'].get(sheet_name, {})
    
    def get_formula_columns(self, sheet_name: str) -> List[int]:
        """获取指定Sheet的公式列"""
        sheet_info = self.template_info['sheets'].get(sheet_name, {})
        return sheet_info.get('formula_columns', [])
    
    def get_header_row(self, sheet_name: str) -> int:
        """获取表头行号"""
        sheet_info = self.template_info['sheets'].get(sheet_name, {})
        return sheet_info.get('header_row', 1)
    
    def get_headers(self, sheet_name: str) -> List[Dict[str, Any]]:
        """获取表头列表"""
        sheet_info = self.template_info['sheets'].get(sheet_name, {})
        return sheet_info.get('headers', [])
    
    def get_template_copy(self):
        """获取模板的深拷贝"""
        return openpyxl.load_workbook(
            self.template_path,
            data_only=False,
            keep_vba=False
        )
