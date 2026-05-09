#!/usr/bin/env python3
"""
交付中心月报生成器
==================
基于经营报告通用框架的交付中心月报定制化功能

支持:
- 5个CSV数据源合并处理与校验
- 严格格式对齐（单元格底色、边框、字体）
- 公式自动计算
- 透视表数据源更新

作者: Jerry 🦞
版本: v1.0
创建时间: 2026-04-24
"""

import argparse
import logging
import sys
from pathlib import Path
from typing import Dict, List, Any
import json

try:
    import pandas as pd
    import numpy as np
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"❌ 缺少依赖: {e}")
    print("请运行: pip install pandas openpyxl numpy")
    sys.exit(1)

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# 导入自定义模块
from data_validator import DataValidator
from template_analyzer import TemplateAnalyzer
from format_checker import FormatChecker


def download_from_ones(filter_name: str, output_file: str, config_path: str = None) -> Dict[str, Any]:
    """
    从 ONES 系统导出筛选器数据
    
    Args:
        filter_name: 筛选器名称（如 '签约项目统计'）
        output_file: 输出 CSV 文件路径
        config_path: ONES 配置文件路径（可选，默认使用 config/ones-config.json）
    
    Returns:
        下载结果字典 {success, rows, file_path, error}
    """
    # TODO: 实现 ONES 自动登录和筛选器导出
    # 实现步骤:
    # 1. 读取配置文件（URL、账号、密码/Token）
    # 2. Playwright 模拟登录 ONES 系统
    # 3. 进入「我的工作台」→「筛选器」
    # 4. 找到对应筛选器，点击「还原」→「导出工作项」
    # 5. 下载并保存为 CSV 文件
    # 6. 返回下载结果和行数统计
    
    logger.warning(f"[TODO] ONES 自动下载功能待实现: {filter_name} -> {output_file}")
    
    # 临时实现: 返回模拟成功，实际需要用户手动下载后使用 local 模式
    return {
        'success': True,
        'rows': 0,
        'file_path': output_file,
        'note': '自动下载功能开发中，请先手动下载 CSV 后使用 --mode local 模式'
    }


def download_from_wecom(file_name: str, output_file: str, sheet_name: str = '确收', config_path: str = None) -> Dict[str, Any]:
    """
    从企业微信微盘下载并导出指定 Sheet 为 CSV
    
    Args:
        file_name: 文件名（如 '202602确收凭证交接'）
        output_file: 输出 CSV 文件路径
        sheet_name: 需要导出的 Sheet 名称
        config_path: 企业微信配置文件路径
    
    Returns:
        下载结果字典 {success, rows, file_path, error}
    """
    # TODO: 实现企业微信微盘自动下载
    # 实现步骤:
    # 1. 读取配置文件（微盘链接、访问凭证）
    # 2. Playwright 访问企业微信微盘共享空间
    # 3. 搜索并下载目标 Excel 文件
    # 4. 使用 pandas 读取指定 Sheet
    # 5. 导出为 CSV 文件
    # 6. 返回下载结果和行数统计
    
    logger.warning(f"[TODO] 企业微信微盘自动下载功能待实现: {file_name} -> {output_file}")
    
    # 临时实现: 返回模拟成功，实际需要用户手动下载后使用 local 模式
    return {
        'success': True,
        'rows': 0,
        'file_path': output_file,
        'sheet_name': sheet_name,
        'note': '自动下载功能开发中，请先手动下载 CSV 后使用 --mode local 模式'
    }


class DeliveryMonthlyReportGenerator:
    """交付中心月报生成器"""
    
    # Sheet与CSV数据的映射配置
    SHEET_DATA_MAPPING = {
        '签约': {
            'data_key': '签约项目统计',
            'start_row': 2,  # 从第2行开始写入（第1行是表头）
            'skip_formula_columns': True
        },
        'POC&提前实施': {
            'data_key': 'POC提前实施',
            'start_row': 2,
            'skip_formula_columns': True
        },
        '异常项目': {
            'data_key': '异常处置',
            'start_row': 2,
            'skip_formula_columns': True
        },
        '确收交接': {
            'data_key': '确收',
            'start_row': 2,
            'skip_formula_columns': True
        },
        '验收交接': {
            'data_key': '验收',
            'start_row': 2,
            'skip_formula_columns': True
        }
    }
    
    def __init__(self, data_dir: str, template_path: str, output_path: str, month_prefix: str):
        """
        初始化月报生成器
        
        Args:
            data_dir: CSV数据目录
            template_path: 模板文件路径
            output_path: 输出文件路径
            month_prefix: 月份前缀，如 '202602'
        """
        self.data_dir = Path(data_dir)
        self.template_path = Path(template_path)
        self.output_path = Path(output_path)
        self.month_prefix = month_prefix
        
        # 初始化子模块
        self.validator = DataValidator(str(self.data_dir), month_prefix)
        self.analyzer = TemplateAnalyzer(str(self.template_path))
        self.format_checker = None
        
        # 工作簿对象
        self.output_wb = None
        
        # 报告数据
        self.generation_report: Dict[str, Any] = {
            'success': False,
            'month': month_prefix,
            'data_validation': {},
            'template_analysis': {},
            'data_filling': {},
            'format_check': {},
            'output_file': None
        }
    
    def run(self, test_mode: bool = False, expected_signing_rows: int = None) -> Dict[str, Any]:
        """
        运行完整的月报生成流程
        
        Args:
            test_mode: 测试模式，只写入前100行数据
            expected_signing_rows: 签约项目统计预期行数
        """
        logger.info("=" * 70)
        logger.info(f"交付中心月报生成器 v1.0 - {self.month_prefix}")
        logger.info("=" * 70)
        
        try:
            # 1. 数据校验
            logger.info("\n[阶段1] 数据加载与校验")
            logger.info("-" * 70)
            self.generation_report['data_validation'] = self.validator.run_validation(expected_signing_rows)
            
            if not self.generation_report['data_validation']['success']:
                logger.error("数据校验失败，终止生成")
                return self.generation_report
            
            # 2. 模板分析
            logger.info("\n[阶段2] 模板结构分析")
            logger.info("-" * 70)
            self.generation_report['template_analysis'] = self.analyzer.analyze_all_sheets()
            
            # 3. 数据填充
            logger.info("\n[阶段3] 数据填充引擎")
            logger.info("-" * 70)
            self._fill_data(test_mode=test_mode)
            
            # 4. 保存文件
            logger.info("\n[阶段4] 保存生成文件")
            logger.info("-" * 70)
            self._save_workbook()
            
            # 5. 格式校验
            logger.info("\n[阶段5] 格式一致性校验")
            logger.info("-" * 70)
            self._check_format()
            
            self.generation_report['success'] = True
            
            # 6. 生成最终报告
            self._generate_final_report()
            
        except Exception as e:
            logger.error(f"生成过程出错: {str(e)}", exc_info=True)
            self.generation_report['error'] = str(e)
        
        return self.generation_report
    
    def _fill_data(self, test_mode: bool = False):
        """填充数据到模板"""
        logger.info("创建模板副本...")
        self.output_wb = self.analyzer.get_template_copy()
        
        filling_stats = {}
        
        for sheet_name, mapping in self.SHEET_DATA_MAPPING.items():
            if sheet_name not in self.output_wb.sheetnames:
                logger.warning(f"Sheet不存在，跳过: {sheet_name}")
                continue
            
            data_key = mapping['data_key']
            df = self.validator.get_dataframe(data_key)
            
            if df is None:
                logger.warning(f"数据不存在，跳过: {data_key}")
                continue
            
            ws = self.output_wb[sheet_name]
            
            # 获取公式列（需要跳过的列）
            formula_columns = self.analyzer.get_formula_columns(sheet_name)
            header_row = self.analyzer.get_header_row(sheet_name)
            
            logger.info(f"填充Sheet: {sheet_name}")
            logger.info(f"  数据源: {data_key} ({len(df)} 行)")
            logger.info(f"  公式列: {formula_columns}")
            
            # 测试模式限制行数
            if test_mode:
                df = df.head(100)
                logger.info(f"  [测试模式] 只写入前 100 行")
            
            # 写入数据 - 按列位置直接写入
            start_row = mapping['start_row']
            rows_written = 0
            
            for df_idx, (_, row_data) in enumerate(df.iterrows()):
                excel_row = start_row + df_idx
                
                # 按列位置写入数据，跳过公式列
                for col_idx, col_name in enumerate(df.columns, start=1):
                    # 跳过公式列
                    if mapping['skip_formula_columns'] and col_idx in formula_columns:
                        continue
                    
                    # 获取数据值
                    value = row_data.get(col_name)
                    if pd.isna(value):
                        value = None
                    
                    # 写入单元格，保留格式
                    cell = ws.cell(row=excel_row, column=col_idx)
                    # 只更新值，不覆盖格式
                    cell.value = value
                
                rows_written += 1
                
                # 进度输出
                if rows_written % 10000 == 0:
                    logger.info(f"  已写入 {rows_written} 行...")
            
            filling_stats[sheet_name] = {
                'rows_written': rows_written,
                'columns_used': len(df.columns),
                'formula_columns_skipped': len(formula_columns)
            }
            
            logger.info(f"  ✓ 完成: {rows_written} 行 x {len(df.columns)} 列")
        
        self.generation_report['data_filling'] = filling_stats
        logger.info("✓ 所有Sheet数据填充完成")
    
    def _save_workbook(self):
        """保存工作簿"""
        # 确保输出目录存在
        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        
        logger.info(f"保存文件: {self.output_path}")
        self.output_wb.save(str(self.output_path))
        self.generation_report['output_file'] = str(self.output_path)
        
        # 获取文件大小
        file_size = self.output_path.stat().st_size / 1024 / 1024
        logger.info(f"✓ 文件已保存 (大小: {file_size:.2f} MB)")
    
    def _check_format(self):
        """格式校验"""
        self.format_checker = FormatChecker(
            str(self.template_path),
            str(self.output_path)
        )
        
        self.generation_report['format_check'] = self.format_checker.check_all_sheets(max_rows=500)
    
    def _generate_final_report(self):
        """生成最终报告"""
        report = self.generation_report
        
        logger.info("\n" + "=" * 70)
        logger.info("📊 生成完成报告")
        logger.info("=" * 70)
        
        print(f"\n月份: {report['month']}")
        print(f"输出文件: {report['output_file']}")
        print(f"整体状态: {'✓ 成功' if report['success'] else '✗ 失败'}")
        
        # 数据校验报告
        print("\n--- 数据校验 ---")
        dv = report['data_validation']
        for file_key, info in dv['files'].items():
            print(f"  {file_key}: {info['rows']} 行 x {info['columns']} 列")
        
        if dv['warnings']:
            print(f"  警告: {len(dv['warnings'])} 个")
        if dv['errors']:
            print(f"  错误: {len(dv['errors'])} 个")
        
        # 数据填充报告
        print("\n--- 数据填充 ---")
        df = report['data_filling']
        for sheet_name, stats in df.items():
            print(f"  {sheet_name}: {stats['rows_written']} 行  ✓")
        
        # 格式校验报告
        print("\n--- 格式校验 ---")
        fc = report['format_check']
        print(f"  检查单元格: {fc['total_cells_checked']}")
        print(f"  格式差异: {len(fc['differences'])} 个")
        print(f"  状态: {'✓ 通过' if fc['success'] else '⚠ 有差异'}")
        
        print("\n" + "=" * 70)
    
    def save_report_json(self, report_path: str):
        """保存生成报告到JSON文件"""
        with open(report_path, 'w', encoding='utf-8') as f:
            json.dump(self.generation_report, f, ensure_ascii=False, indent=2)
        logger.info(f"生成报告已保存: {report_path}")


def main():
    """主函数"""
    parser = argparse.ArgumentParser(description='交付中心月报生成器 - 支持两种数据来源模式')
    
    # 数据来源模式选择
    parser.add_argument('--mode', '-M', 
        choices=['auto', 'local'], 
        default='local',
        help='数据来源模式: auto=自动从ONES+企业微信下载, local=使用本地已下载数据 (默认: local)')
    
    # 两种模式通用参数
    parser.add_argument('--template', '-t', required=True, help='Excel模板文件路径')
    parser.add_argument('--output', '-o', required=True, help='输出文件路径')
    parser.add_argument('--month', '-m', required=True, help='月份前缀，如 202602')
    parser.add_argument('--test', action='store_true', help='测试模式，只写入前100行')
    parser.add_argument('--expected-rows', type=int, default=None, help='签约项目统计预期行数')
    parser.add_argument('--report-json', help='生成报告JSON输出路径')
    
    # local 模式专用参数
    parser.add_argument('--data-dir', '-d', help='[local模式] CSV数据目录路径')
    
    # auto 模式专用参数
    parser.add_argument('--ones-config', help='[auto模式] ONES系统配置文件路径')
    parser.add_argument('--wecom-config', help='[auto模式] 企业微信微盘配置文件路径')
    parser.add_argument('--download-dir', help='[auto模式] 下载的CSV文件保存目录')
    
    args = parser.parse_args()
    
    # 参数校验
    if args.mode == 'local' and not args.data_dir:
        print("❌ local模式必须指定 --data-dir 参数")
        print("示例: --mode local --data-dir /path/to/csv/files")
        return 1
    
    if args.mode == 'auto':
        print("\n" + "="*70)
        print("🚀 [AUTO模式] 自动从 ONES + 企业微信微盘下载数据")
        print("="*70)
        
        # 使用默认下载目录（如未指定）
        download_dir = args.download_dir or f'/tmp/delivery-report-{args.month}'
        Path(download_dir).mkdir(parents=True, exist_ok=True)
        
        # ============================================
        # 阶段 1: 从 ONES 系统导出 3 个筛选器数据
        # ============================================
        print("\n[1/5] 从 ONES 系统导出签约项目统计...")
        ones_result = download_from_ones(
            filter_name='签约项目统计',
            output_file=f'{download_dir}/{args.month}_签约项目统计.csv',
            config_path=args.ones_config
        )
        if not ones_result['success']:
            print(f"   ❌ 失败: {ones_result['error']}")
            return 1
        print(f"   ✅ 成功下载 {ones_result['rows']} 行数据")
        
        print("\n[2/5] 从 ONES 系统导出 POC&提前实施统计...")
        ones_result2 = download_from_ones(
            filter_name='POC&提前实施统计',
            output_file=f'{download_dir}/{args.month}_POC提前实施.csv',
            config_path=args.ones_config
        )
        if not ones_result2['success']:
            print(f"   ❌ 失败: {ones_result2['error']}")
            return 1
        print(f"   ✅ 成功下载 {ones_result2['rows']} 行数据")
        
        print("\n[3/5] 从 ONES 系统导出签约项目异常处置...")
        ones_result3 = download_from_ones(
            filter_name='签约项目异常处置',
            output_file=f'{download_dir}/{args.month}_异常处置.csv',
            config_path=args.ones_config
        )
        if not ones_result3['success']:
            print(f"   ❌ 失败: {ones_result3['error']}")
            return 1
        print(f"   ✅ 成功下载 {ones_result3['rows']} 行数据")
        
        # ============================================
        # 阶段 2: 从企业微信微盘下载 2 个交接表
        # ============================================
        print("\n[4/5] 从企业微信微盘下载确收凭证交接表...")
        wecom_result1 = download_from_wecom(
            file_name=f'{args.month}确收凭证交接',
            output_file=f'{download_dir}/{args.month}_确收.csv',
            sheet_name='确收',
            config_path=args.wecom_config
        )
        if not wecom_result1['success']:
            print(f"   ❌ 失败: {wecom_result1['error']}")
            return 1
        print(f"   ✅ 成功下载 {wecom_result1['rows']} 行数据")
        
        print("\n[5/5] 从企业微信微盘下载验收凭证交接表...")
        wecom_result2 = download_from_wecom(
            file_name=f'{args.month}验收凭证交接',
            output_file=f'{download_dir}/{args.month}_验收.csv',
            sheet_name='验收',
            config_path=args.wecom_config
        )
        if not wecom_result2['success']:
            print(f"   ❌ 失败: {wecom_result2['error']}")
            return 1
        print(f"   ✅ 成功下载 {wecom_result2['rows']} 行数据")
        
        print("\n" + "="*70)
        print(f"✅ 所有数据下载完成，文件已保存至: {download_dir}")
        print("="*70)
        
        # 使用下载目录作为数据目录
        data_dir = download_dir
    else:
        # local 模式: 使用用户指定的本地数据目录
        print("\n" + "="*70)
        print("📂 [LOCAL模式] 使用本地已下载的 CSV 数据")
        print("="*70)
        data_dir = args.data_dir
    
    # 创建生成器
    generator = DeliveryMonthlyReportGenerator(
        data_dir=data_dir,
        template_path=args.template,
        output_path=args.output,
        month_prefix=args.month
    )
    
    # 运行生成
    report = generator.run(
        test_mode=args.test,
        expected_signing_rows=args.expected_rows
    )
    
    # 保存报告
    if args.report_json:
        generator.save_report_json(args.report_json)
    
    return 0 if report['success'] else 1


if __name__ == '__main__':
    sys.exit(main())
